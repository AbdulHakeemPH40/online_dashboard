from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib import messages
from django.core.cache import cache
from django.views.decorators.http import require_http_methods
from .models import Outlet, Item, ItemOutlet
from .utils import decode_csv_upload
from .promotion_service import PromotionService
import logging
from decimal import Decimal, InvalidOperation
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from functools import wraps
from datetime import datetime, timedelta, date
import json

logger = logging.getLogger(__name__)

# Allowed wrap values for items
ALLOWED_WRAP_VALUES = {"9900", "10000"}


def rate_limit(max_requests: int, time_window_seconds: int):
    """
    Simple rate limiting decorator using Django cache.
    Limits number of requests per user per time window.
    
    Args:
        max_requests: Maximum number of requests allowed
        time_window_seconds: Time window in seconds
    
    Returns:
        HttpResponse with 429 status if limit exceeded
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            # Get user identifier
            user_key = f"rate_limit:{view_func.__name__}:{request.user.id or request.META.get('REMOTE_ADDR')}"
            
            # Get current count
            current_count = cache.get(user_key, 0)
            
            if current_count >= max_requests:
                logger.warning(f"Rate limit exceeded for {request.user} on {view_func.__name__}")
                return JsonResponse({
                    'success': False,
                    'message': f'Rate limit exceeded. Maximum {max_requests} requests per {time_window_seconds} seconds.'
                }, status=429)
            
            # Increment counter
            cache.set(user_key, current_count + 1, time_window_seconds)
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def login_view(request):
    """
    Custom login view that handles user authentication
    """
    if request.method == 'GET':
        # If user is already authenticated, redirect to dashboard
        if request.user.is_authenticated:
            return redirect('dashboard')
        return render(request, 'login.html')
    
    elif request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect to dashboard after successful login
                next_url = request.GET.get('next', 'dashboard')
                return redirect(next_url)
            else:
                # Invalid credentials
                return render(request, 'login.html', {
                    'error': 'Invalid username or password. Please try again.'
                })
        else:
            # Missing username or password
            return render(request, 'login.html', {
                'error': 'Please enter both username and password.'
            })


@login_required
def dashboard(request):
    """
    Main dashboard view that renders the dashboard template for Pasons
    OPTIMIZED: Uses simple database counts instead of complex Python iteration
    """
    try:
        from django.db.models import Sum, Count, Q
        
        # Count total items for this platform
        pasons_total_items = Item.objects.filter(platform='pasons', is_active=True).count()
        
        # Count items that have at least one ItemOutlet with stock > 0
        # This is a simple approximation - active = has stock
        items_with_stock = ItemOutlet.objects.filter(
            outlet__platforms='pasons',
            item__platform='pasons',
            outlet__is_active=True,
            outlet_stock__gt=0
        ).values('item_id').distinct().count()
        
        # Items with outlets but zero/low stock
        items_with_outlets = ItemOutlet.objects.filter(
            outlet__platforms='pasons',
            item__platform='pasons',
            outlet__is_active=True
        ).values('item_id').distinct().count()
        
        pasons_active_items = items_with_stock
        pasons_low_stock_items = max(0, items_with_outlets - items_with_stock)
        
    except Exception as e:
        import sys
        print(f'Dashboard error: {str(e)}', file=sys.stderr)
        pasons_total_items = 0
        pasons_active_items = 0
        pasons_low_stock_items = 0

    context = {
        'page_title': 'Pasons Dashboard Overview',
        'active_nav': 'pasons',
        'total_items': pasons_total_items,
        'active_items': pasons_active_items,
        'low_stock_items': pasons_low_stock_items,
    }
    return render(request, 'dashboard.html', context)


@login_required
def talabat_dashboard(request):
    """
    Dashboard view that renders the Talabat dashboard template
    OPTIMIZED: Uses simple database counts instead of complex Python iteration
    """
    try:
        from django.db.models import Sum, Count, Q
        
        # Count total items for this platform
        talabat_total_items = Item.objects.filter(platform='talabat', is_active=True).count()
        
        # Count items that have at least one ItemOutlet with stock > 0
        items_with_stock = ItemOutlet.objects.filter(
            outlet__platforms='talabat',
            item__platform='talabat',
            outlet__is_active=True,
            outlet_stock__gt=0
        ).values('item_id').distinct().count()
        
        # Items with outlets but zero/low stock
        items_with_outlets = ItemOutlet.objects.filter(
            outlet__platforms='talabat',
            item__platform='talabat',
            outlet__is_active=True
        ).values('item_id').distinct().count()
        
        talabat_active_items = items_with_stock
        talabat_low_stock_items = max(0, items_with_outlets - items_with_stock)
        
    except Exception as e:
        import sys
        print(f'Dashboard error: {str(e)}', file=sys.stderr)
        talabat_total_items = 0
        talabat_active_items = 0
        talabat_low_stock_items = 0

    context = {
        'page_title': 'Talabat Dashboard Overview',
        'active_nav': 'talabat',
        'total_items': talabat_total_items,
        'active_items': talabat_active_items,
        'low_stock_items': talabat_low_stock_items,
    }
    return render(request, 'talabat_dashboard.html', context)


@login_required
def erp_page(request):
    """
    ERP Integration page for Talabat platform.
    Export data with format: Party, Item Code, Location, Unit, Price
    """
    from .models import ERPExportHistory
    
    try:
        # Get Talabat outlets for dropdown
        talabat_outlets = Outlet.objects.filter(platforms='talabat', is_active=True).order_by('name')
        
        # Get ERP export history
        erp_exports = ERPExportHistory.objects.all().order_by('-export_timestamp')[:100]
        
    except Exception as e:
        logger.error(f"ERP page error: {str(e)}")
        talabat_outlets = []
        erp_exports = []
    
    context = {
        'page_title': 'ERP Integration - Talabat',
        'active_nav': 'erp',
        'talabat_outlets': talabat_outlets,
        'erp_exports': erp_exports,
    }
    return render(request, 'erp_page.html', context)


@login_required
def list_items_api(request):
    """
    API endpoint to list items filtered by platform with optional search and pagination.
    Returns data suitable for populating the dashboard table.
    OPTIMIZED: Uses database-level annotations to avoid N+1 queries.
    """
    try:
        platform = request.GET.get('platform', '').strip()
        if platform not in ('pasons', 'talabat'):
            return JsonResponse({'success': False, 'message': 'Invalid or missing platform'}, status=400)

        query = request.GET.get('q', '').strip()
        page = int(request.GET.get('page', '1'))
        page_size = int(request.GET.get('page_size', '50'))
        if page_size <= 0:
            page_size = 50
        if page_size > 200:
            page_size = 200

        from django.db.models import Sum, OuterRef, Subquery, Value
        from django.db.models.functions import Coalesce
        
        # Build base queryset with platform filter
        qs = Item.objects.filter(platform=platform)

        if query:
            qs = qs.filter(
                Q(item_code__icontains=query) |
                Q(description__icontains=query) |
                Q(sku__icontains=query) |
                Q(barcode__icontains=query) |
                Q(pack_description__icontains=query)
            )

        # OPTIMIZATION: Annotate stock sum at database level (single query)
        # This adds a 'platform_stock' field to each item via subquery
        stock_subquery = ItemOutlet.objects.filter(
            item=OuterRef('pk'),
            outlet__platforms=platform,
            outlet__is_active=True
        ).values('item').annotate(
            total=Sum('outlet_stock')
        ).values('total')
        
        qs = qs.annotate(
            platform_stock=Coalesce(Subquery(stock_subquery), Value(0))
        )

        paginator = Paginator(qs.order_by('item_code'), page_size)
        page_obj = paginator.get_page(page)
        
        # Get item IDs for batch outlet lookup
        item_ids = [item.id for item in page_obj.object_list]
        
        # OPTIMIZATION: Batch fetch all outlet names in ONE query
        # Include store_id to prevent duplicate names when multiple outlets have same name
        outlet_data = ItemOutlet.objects.filter(
            item_id__in=item_ids,
            outlet__platforms=platform,
            outlet__is_active=True
        ).select_related('outlet').values('item_id', 'outlet__name', 'outlet__store_id')
        
        # Build lookup: item_id -> set of outlet display names (with store_id for uniqueness)
        outlet_names_map = {}
        for od in outlet_data:
            if od['item_id'] not in outlet_names_map:
                outlet_names_map[od['item_id']] = set()
            # Include store_id to differentiate outlets with identical names
            outlet_display = od['outlet__name']
            if od['outlet__store_id']:
                outlet_display = f"{outlet_display} ({od['outlet__store_id']})"
            outlet_names_map[od['item_id']].add(outlet_display)

        items = []
        for item in page_obj.object_list:
            # Sort outlet names for consistent, readable display
            outlet_names = sorted(list(outlet_names_map.get(item.id, [])))
            
            items.append({
                'item_code': item.item_code,
                'description': item.description,
                'pack_description': item.pack_description or '',
                'locations': outlet_names,
                'mrp': float(item.mrp),
                'stock': int(item.platform_stock or 0),  # From annotation
                'status': 'Active' if item.is_active else 'Inactive',
                'status_code': 'A' if item.is_active else 'I',
                'lock_status': 'N/A',
                'cost': float(item.cost),
                'sku': item.sku,
                'talabat_margin': float(item.effective_talabat_margin) if platform == 'talabat' and item.effective_talabat_margin is not None else None,
            })

        return JsonResponse({
            'success': True,
            'items': items,
            'page': page_obj.number,
            'page_size': page_size,
            'total_items': paginator.count,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        })
    except Exception as e:
        logger.error(f"list_items_api error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Error loading items: {str(e)}'})

def dashboard_stats_api(request):
    """
    API endpoint that provides dashboard statistics in JSON format
    """
    # Placeholder data for fresh start
    return JsonResponse({
        'total_products': 0,
        'active_products': 0,
        'low_stock_products': 0,
        'out_of_stock_products': 0,
        'platform_mapping': {
            'talabat': 0,
            'pasons': 0
        },
        'last_sync': {
            'talabat': None,
            'pasons': None
        },
        'category_breakdown': {},
        'stock_status': {
            'in_stock': 0,
            'low_stock': 0,
            'out_of_stock': 0
        },
        'recent_syncs': [],
        'alerts': [],
        'price_stats': {
            'min_price': 0,
            'max_price': 0,
            'avg_price': 0
        },
        'active_pricing_rules': 0,
        'weight_variants_count': 0,
        'message': 'Dashboard ready for fresh implementation'
    })


def system_health_check(request):
    """
    API endpoint for system health monitoring
    """
    return JsonResponse({
        'status': 'healthy',
        'components': {
            'database': 'healthy',
            'sync_operations': 'healthy',
            'error_rate': 'healthy'
        },
        'metrics': {
            'recent_errors': 0,
            'last_successful_sync': True
        },
        'message': 'System ready for fresh implementation'
    })


def quick_stats(request):
    """
    Lightweight API endpoint for quick dashboard updates
    """
    return JsonResponse({
        'total_products': 0,
        'active_products': 0,
        'running_syncs': 0,
        'message': 'Ready for implementation'
    })


@login_required
def create_store(request):
    """Create new store/outlet"""
    if request.method == 'POST':
        name = request.POST.get('name')
        location = request.POST.get('location')
        platforms = request.POST.get('platforms')
        
        if name and location and platforms:
            # Check if store name already exists for this platform
            if Outlet.objects.filter(name=name, platforms=platforms).exists():
                platform_name = {
                    'pasons': 'Pasons Ecommerce',
                    'talabat': 'Talabat Platform'
                }.get(platforms, platforms)
                messages.error(request, f'Store "{name}" already exists on {platform_name}. Please use a different name.')
            else:
                try:
                    outlet = Outlet.objects.create(
                        name=name,
                        location=location,
                        platforms=platforms,
                        is_active=True
                    )
                    platform_name = {
                        'pasons': 'Pasons Ecommerce',
                        'talabat': 'Talabat Platform'
                    }.get(platforms, platforms)
                    
                    messages.success(request, f'Store "{name}" created successfully with ID: {outlet.store_id} for {platform_name}')
                    return redirect('integration:store_list')
                except Exception as e:
                    messages.error(request, f'Error creating store: {str(e)}')
        else:
            messages.error(request, 'Please fill in all required fields including platform selection.')
    
    context = {
        'page_title': 'Create New Store',
        'active_nav': 'stores'
    }
    return render(request, 'create_store.html', context)


@login_required
def store_list(request):
    """List all stores"""
    outlets = Outlet.objects.all().order_by('-created_at')
    
    # Calculate platform-specific statistics
    pasons_outlets = outlets.filter(platforms='pasons')
    talabat_outlets = outlets.filter(platforms='talabat')
    
    context = {
        'page_title': 'Store Management',
        'active_nav': 'stores',
        'outlets': outlets,
        'total_stores': outlets.count(),
        'active_stores': outlets.filter(is_active=True).count(),
        'pasons_stores': pasons_outlets.count(),
        'talabat_stores': talabat_outlets.count()
    }
    return render(request, 'store_list.html', context)


@login_required
def edit_store(request, store_id):
    """Edit existing store/outlet"""
    from .models import Outlet
    from django.shortcuts import get_object_or_404
    
    outlet = get_object_or_404(Outlet, id=store_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        location = request.POST.get('location')
        platforms = request.POST.get('platforms')
        
        if name and location and platforms:
            # Check if store name already exists for this platform (excluding current store)
            if Outlet.objects.filter(name=name, platforms=platforms).exclude(id=outlet.id).exists():
                platform_name = {
                    'pasons': 'Pasons Ecommerce',
                    'talabat': 'Talabat Platform'
                }.get(platforms, platforms)
                messages.error(request, f'Store "{name}" already exists on {platform_name}. Please use a different name.')
            else:
                # Check if platform has changed
                platform_changed = outlet.platforms != platforms
                
                outlet.name = name
                outlet.location = location
                
                # If platform changed, generate new store ID
                if platform_changed:
                    outlet.platforms = platforms
                    # Generate new store ID based on new platform
                    outlet.store_id = outlet.generate_unique_store_id()
                    messages.success(request, f'Store updated successfully! New Store ID: {outlet.store_id} (Platform: {platforms.title()})')
                else:
                    messages.success(request, f'Store "{name}" updated successfully!')
                
                outlet.save()
                return redirect('integration:store_list')
        else:
            messages.error(request, 'Please fill in all required fields.')
    
    context = {
        'page_title': 'Edit Store',
        'active_nav': 'stores',
        'outlet': outlet
    }
    return render(request, 'edit_store.html', context)


@login_required
def delete_store(request, store_id):
    """Delete existing store/outlet"""
    from .models import Outlet
    from django.shortcuts import get_object_or_404
    
    outlet = get_object_or_404(Outlet, id=store_id)
    
    if request.method == 'POST':
        outlet.delete()
        return redirect('integration:store_list')
    
    context = {
        'page_title': 'Delete Store',
        'active_nav': 'stores',
        'outlet': outlet
    }
    return render(request, 'delete_store.html', context)


@login_required
def bulk_item_creation(request):
    """
    Bulk item creation view with platform-specific outlet filtering
    """
    if request.method == 'POST':
        # Handle CSV upload and item creation
        platform = request.POST.get('platform')
        csv_file = request.FILES.get('csv_file')
        
        if platform and csv_file:
            try:
                # Get all outlets that support the selected platform
                from .models import Outlet, Item, ItemOutlet
                from django.contrib import messages
                from django.db import models, transaction
                
                # Filter outlets by platform - STRICT ISOLATION
                outlets = Outlet.objects.filter(
                    is_active=True,
                    platforms=platform
                )
                
                if not outlets.exists():
                    messages.error(request, f"No active outlets found for {platform.title()} platform.")
                    return redirect('integration:bulk_item_creation')
                
                # Process CSV file
                import csv
                import io

                # Read CSV content with encoding fallback
                csv_content, _encoding_used = decode_csv_upload(csv_file)
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                # Strict header validation
                required_headers = {'wrap', 'item_code', 'description', 'units', 'sku', 'pack_description'}
                optional_headers = {'barcode', 'mrp', 'selling_price', 'cost', 'stock', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty'}
                allowed_headers = required_headers | optional_headers
                # Filter out empty header fields (from trailing delimiters)
                # Use normalize_csv_header for proper BOM/invisible char handling
                from .utils import normalize_csv_header
                headers = [normalize_csv_header(h) for h in (csv_reader.fieldnames or []) if h and h.strip()]
                if not headers:
                    messages.error(request, 'CSV file is missing a header row. Include headers exactly as specified.')
                    return redirect('integration:bulk_item_creation')
                # Disallow is_active column entirely
                if 'is_active' in headers:
                    messages.error(request, "Column 'is_active' is not allowed in bulk creation CSV. Please remove it.")
                    return redirect('integration:bulk_item_creation')
                missing_required = sorted(list(required_headers - set(headers)))
                unknown_headers = sorted([h for h in headers if h and h not in allowed_headers])
                if missing_required:
                    messages.error(request, f"Missing required columns: {', '.join(missing_required)}. The file was rejected.")
                    return redirect('integration:bulk_item_creation')
                if unknown_headers:
                    messages.error(request, f"Unknown columns present: {', '.join(unknown_headers)}. Only defined headers are allowed. The file was rejected.")
                    return redirect('integration:bulk_item_creation')
                
                errors = []
                BATCH_SIZE = 1000
                
                # Collect one unique row per SKU from CSV (deduplicate within upload)
                unique_rows_by_sku = {}
                duplicate_skus_in_csv = []
                # OPTIMIZATION: Track wrap=10000 items by (item_code, units) for O(1) duplicate check
                wrap10000_pairs_in_csv = {}  # {(item_code, units): sku}
                
                for row_num, original_row in enumerate(csv_reader, start=2):  # Start from 2 (header is row 1)
                    try:
                        # Normalize row keys to lowercase for consistent access
                        row = {k.lower().strip(): v for k, v in original_row.items()}
                        
                        base_item_code = row.get('item_code', '').strip()
                        base_sku = row.get('sku', '').strip()
                        
                        # Validate mandatory fields first
                        mandatory_fields = {
                            'wrap': row.get('wrap', '').strip(),
                            'item_code': base_item_code,
                            'description': row.get('description', '').strip(),
                            'units': row.get('units', '').strip(),
                            'sku': base_sku,
                            'pack_description': row.get('pack_description', '').strip()
                        }
                        
                        missing_fields = [field for field, value in mandatory_fields.items() if not value]
                        if missing_fields:
                            errors.append(f"Row {row_num}: Missing mandatory fields: {', '.join(missing_fields)}")
                            continue

                        # Validate wrap strictly
                        if mandatory_fields['wrap'] not in ALLOWED_WRAP_VALUES:
                            errors.append(f"Row {row_num}: Wrap must be 9900 or 10000 (got '{mandatory_fields['wrap']}')")
                            continue
                        
                        # No additional mandatory field validation needed

                        # Optional fields parsing
                        barcode = row.get('barcode', '').strip()
                        # NOTE: mrp, selling_price, cost, stock are OUTLET-SPECIFIC
                        # They should NOT be set during item creation (always default to 0)
                        weight_division_factor_str = row.get('weight_division_factor', '').strip()
                        outer_case_quantity_str = row.get('outer_case_quantity', '').strip()
                        minimum_qty_str = row.get('minimum_qty', '').strip()

                        # Defaults (outlet-specific fields always 0)
                        weight_division_factor = None
                        outer_case_quantity = None
                        minimum_qty = None

                        # Numeric validations for optional fields
                        try:
                            if weight_division_factor_str:
                                weight_division_factor = Decimal(weight_division_factor_str)
                            if outer_case_quantity_str:
                                outer_case_quantity = int(outer_case_quantity_str)
                            if minimum_qty_str:
                                minimum_qty = int(minimum_qty_str)
                        except (InvalidOperation, ValueError):
                            errors.append(f"Row {row_num}: Invalid numeric values in one of [weight_division_factor, outer_case_quantity, minimum_qty]")
                            continue

                        # No boolean parsing; is_active not allowed in CSV
                        
                        # WRAP=10000 SPECIFIC: Check for duplicates by (item_code, units) within CSV
                        # OPTIMIZATION: Use dictionary lookup O(1) instead of loop O(n)
                        is_duplicate_in_csv = False
                        if mandatory_fields['wrap'] == '10000':
                            pair = (base_item_code, mandatory_fields['units'])
                            if pair in wrap10000_pairs_in_csv:
                                # This pair already exists in CSV with a different SKU
                                if base_sku not in duplicate_skus_in_csv:
                                    duplicate_skus_in_csv.append(base_sku)
                                is_duplicate_in_csv = True
                        
                        # Skip this row if it's a wrap=10000 duplicate
                        if is_duplicate_in_csv:
                            continue
                        
                        # Deduplicate within the uploaded file (by SKU)
                        if base_sku in unique_rows_by_sku:
                            if base_sku not in duplicate_skus_in_csv:
                                duplicate_skus_in_csv.append(base_sku)
                            continue
                        
                        # Store normalized data for later bulk_create
                        unique_rows_by_sku[base_sku] = {
                            'platform': platform,  # Platform isolation
                            'item_code': base_item_code,
                            'description': mandatory_fields['description'],
                            'pack_description': mandatory_fields['pack_description'],
                            'units': mandatory_fields['units'],
                            'sku': base_sku,
                            'barcode': barcode,
                            'wrap': mandatory_fields['wrap'],
                            # Item configuration fields
                            'weight_division_factor': weight_division_factor,
                            'outer_case_quantity': outer_case_quantity,
                            'minimum_qty': minimum_qty,
                            # Outlet-specific fields (always default to 0)
                            'selling_price': 0,
                            'stock': 0,
                            'cost': 0,
                            'mrp': 0,
                        }
                        
                        # Track wrap=10000 pairs for future duplicate checking
                        if mandatory_fields['wrap'] == '10000':
                            pair = (base_item_code, mandatory_fields['units'])
                            wrap10000_pairs_in_csv[pair] = base_sku
                    except Exception as e:
                        logger.error(f"Bulk item creation error - Row {row_num}: {type(e).__name__}: {str(e)}", exc_info=True)
                        errors.append(f"Row {row_num}: Error parsing row - {str(e)}")

                unique_skus = list(unique_rows_by_sku.keys())
                # Items are identified by platform + SKU (unique per platform)
                # Same SKU on different platforms = TWO separate Item records
                # WRAP=10000: Also check for duplicates by (item_code, units) in DATABASE
                
                # Find items that already exist for THIS platform (by platform + SKU)
                existing_items_for_platform = Item.objects.filter(
                    platform=platform,
                    sku__in=unique_skus
                )
                
                # Map by SKU
                existing_items_map = {item.sku: item for item in existing_items_for_platform}
                
                # For wrap=10000 items: Check if (item_code, units) already exists in database
                # OPTIMIZATION: Use simpler approach - avoid massive OR conditions
                wrap10000_duplicates_in_db = []
                
                # Collect all wrap=10000 items to check
                wrap10000_items_to_check = [
                    (sku, row_data) for sku, row_data in unique_rows_by_sku.items()
                    if row_data.get('wrap') == '10000' and sku not in existing_items_map
                ]
                
                # ONLY check database if there are wrap=10000 items in CSV
                if wrap10000_items_to_check:
                    # Get ALL wrap=10000 items for this platform in one simple query (no OR conditions)
                    all_wrap10000_in_db = Item.objects.filter(
                        wrap='10000',
                        platform=platform
                    ).values_list('item_code', 'units')
                    
                    existing_pairs = set(all_wrap10000_in_db)
                    
                    # Mark SKUs as duplicates if their (item_code, units) pair exists in database
                    for sku, row_data in wrap10000_items_to_check:
                        pair = (row_data.get('item_code'), row_data.get('units'))
                        if pair in existing_pairs:
                            wrap10000_duplicates_in_db.append(sku)
                            duplicate_skus_in_csv.append(sku)
                
                # Items to CREATE: SKU doesn't exist on THIS platform AND not a wrap=10000 duplicate
                # Items to UPDATE: SKU exists on THIS platform
                to_create_skus = [sku for sku in unique_skus if sku not in existing_items_map and sku not in wrap10000_duplicates_in_db]
                
                items_to_create = [
                    Item(**unique_rows_by_sku[sku])
                    for sku in to_create_skus
                ]
                
                # If any errors were detected, reject entire file without creating
                if errors:
                    for error in errors[:5]:  # Show first 5 errors
                        messages.error(request, error)
                    if len(errors) > 5:
                        messages.warning(request, f"And {len(errors) - 5} more errors...")
                    # Persist a structured breakdown to show on page
                    try:
                        request.session['bulk_creation_summary'] = {
                            'platform': platform,
                            'outlet_count': outlets.count(),
                            'outlet_names': [o.name for o in outlets],
                            'created_count': 0,
                            'existing_count': 0,
                            'associations_attempted': 0,
                            'associations_created_new_items': 0,
                            'unique_csv_skus': len(unique_skus),
                            'duplicate_skus': duplicate_skus_in_csv,
                            'error_messages': errors,
                            'upload_file_name': getattr(csv_file, 'name', ''),
                        }
                    except Exception:
                        pass
                    messages.error(request, 'CSV validation failed; the file was rejected. Fix the reported issues and try again.')
                    return redirect('integration:bulk_item_creation')

                # Create items in bulk for THIS platform
                created_items_qs = []
                updated_items_count = 0
                with transaction.atomic():
                    if items_to_create:
                        # Get SKUs being created (unique identifiers for this batch)
                        skus_to_create = to_create_skus
                        
                        Item.objects.bulk_create(items_to_create, batch_size=BATCH_SIZE)
                        
                        # Refetch created items by platform + SKU (not just SKU)
                        created_items_qs = list(Item.objects.filter(platform=platform, sku__in=skus_to_create))
                    
                    # UPDATE existing items on THIS platform with new data from CSV
                    for sku, item in existing_items_map.items():
                        csv_data = unique_rows_by_sku.get(sku)
                        if csv_data:
                            # Update fields with new data (only if CSV has value)
                            if csv_data.get('barcode'):
                                item.barcode = csv_data['barcode']
                            if csv_data.get('units'):
                                item.units = csv_data['units']
                            if csv_data.get('description'):
                                item.description = csv_data['description']
                            if csv_data.get('pack_description'):
                                item.pack_description = csv_data['pack_description']
                            if csv_data.get('wrap'):
                                item.wrap = csv_data['wrap']
                            if csv_data.get('weight_division_factor') is not None:
                                item.weight_division_factor = csv_data['weight_division_factor']
                            if csv_data.get('outer_case_quantity') is not None:
                                item.outer_case_quantity = csv_data['outer_case_quantity']
                            if csv_data.get('minimum_qty') is not None:
                                item.minimum_qty = csv_data['minimum_qty']
                            item.save()
                            updated_items_count += 1
                
                # NOTE: Item creation NO LONGER auto-assigns items to outlets
                # Items are only assigned to outlets when price/stock is updated via:
                # - /price-update/
                # - /stock-update/
                # - /product-update/
                
                total_items_created = len(created_items_qs)
                
                # Build clear success message
                if total_items_created > 0:
                    messages.success(request, f"Successfully created {total_items_created} items for {platform.title()} platform.")
                
                if updated_items_count > 0:
                    messages.info(request, f"Updated {updated_items_count} existing items with new data.")
                
                if duplicate_skus_in_csv:
                    if len(duplicate_skus_in_csv) <= 5:
                        messages.info(request, f"Duplicate SKUs in CSV ignored: {', '.join(duplicate_skus_in_csv)}")
                    else:
                        messages.info(request, f"Duplicate SKUs in CSV (first 5): {', '.join(duplicate_skus_in_csv[:5])}... and {len(duplicate_skus_in_csv) - 5} more.")
                
                # Show warning only if truly nothing was processed
                if total_items_created == 0 and updated_items_count == 0 and not errors:
                    messages.warning(request, "No items were processed. Please check your CSV file format and content.")
                
                # Build detailed duplicate information
                # Group duplicates by (item_code, units) combination
                duplicates_detailed = {}
                for sku in duplicate_skus_in_csv:
                    row_data = unique_rows_by_sku.get(sku)
                    if row_data:
                        key = f"{row_data.get('item_code')}|{row_data.get('units')}"
                        if key not in duplicates_detailed:
                            duplicates_detailed[key] = {
                                'item_code': row_data.get('item_code'),
                                'units': row_data.get('units'),
                                'skipped_skus': [],
                                'wrap_type': row_data.get('wrap')
                            }
                        duplicates_detailed[key]['skipped_skus'].append({
                            'sku': sku,
                            'description': row_data.get('description', ''),
                            'reason': 'wrap=10000 duplicate - first occurrence kept'
                        })
                                
                # Convert to list for JSON serialization
                duplicates_list = list(duplicates_detailed.values())
                                
                # At this point, no errors remain; creation has succeeded or associated
                
                # Persist a structured breakdown for the next page render
                try:
                    request.session['bulk_creation_summary'] = {
                        'platform': platform,
                        'created_count': total_items_created,
                        'updated_count': updated_items_count,
                        'existing_count': len(existing_items_map),
                        'unique_csv_skus': len(unique_skus),
                        'duplicate_skus': duplicate_skus_in_csv,
                        'duplicates_detailed': duplicates_list,  # NEW: Detailed duplicate info
                        'error_messages': errors,
                        'upload_file_name': getattr(csv_file, 'name', ''),
                    }
                except Exception:
                    # If session write fails for any reason, continue with redirect
                    pass
                
                # Log upload history
                from .models import UploadHistory
                total_records = len(unique_skus)
                upload_status = 'success' if not errors else ('partial' if total_items_created > 0 else 'failed')
                UploadHistory.objects.create(
                    file_name=csv_file.name,
                    platform=platform,
                    outlet=None,  # Bulk creation is global
                    update_type='bulk_creation',
                    records_total=total_records,
                    records_success=total_items_created + updated_items_count,
                    records_failed=len(errors),
                    records_skipped=len(existing_items_map),
                    status=upload_status,
                    uploaded_by=request.user if request.user.is_authenticated else None,
                )

                return redirect('integration:bulk_item_creation')
                
            except Exception as e:
                logger.error(f"Bulk item creation CSV processing failed: {type(e).__name__}: {str(e)}", exc_info=True)
                messages.error(request, f"Error processing CSV file: {str(e)}")
        else:
            from django.contrib import messages
            if not platform:
                messages.error(request, "Please select a platform.")
            elif not csv_file:
                messages.error(request, "Please select a CSV file.")
            else:
                messages.error(request, "Please fill all required fields and select a CSV file.")
    
    # Get all active outlets for initial load (will be filtered by JavaScript)
    from .models import Outlet
    outlets = Outlet.objects.filter(is_active=True).order_by('name')
    
    # Pull structured breakdown from session (if available)
    bulk_summary = request.session.pop('bulk_creation_summary', None)

    context = {
        'page_title': 'Bulk Item Creation',
        'active_nav': 'bulk_operations',
        'outlets': outlets,
        'bulk_summary': bulk_summary
    }
    return render(request, 'bulk_item_creation.html', context)


@login_required
def get_outlets_by_platform(request):
    """
    AJAX endpoint to get outlets filtered by platform
    """
    platform = request.GET.get('platform')
    
    if not platform:
        return JsonResponse({'success': True, 'outlets': []})
    
    from .models import Outlet
    from django.db import models
    
    # Filter outlets based on platform
    if platform == 'all':
        # Return all active outlets when 'all' is selected
        outlets = Outlet.objects.filter(is_active=True).order_by('name')
    elif platform in ['pasons', 'talabat']:
        # STRICT platform isolation - no 'both'
        outlets = Outlet.objects.filter(
            is_active=True,
            platforms=platform
        ).order_by('name')
    else:
        # Return empty list for invalid platforms
        return JsonResponse({'success': True, 'outlets': []})
    
    outlets_data = []
    for outlet in outlets:
        outlets_data.append({
            'id': outlet.id,
            'name': outlet.name,
            'store_id': outlet.store_id,
            'location': outlet.location,
            'platforms': outlet.platforms
        })
    
    return JsonResponse({'success': True, 'outlets': outlets_data})


@login_required
@rate_limit(max_requests=20, time_window_seconds=60)
def product_update(request):
    """
    Product update view for updating MRP, Cost, and Stock via CSV.
    Uses Item Code + Units combination as unique identifier within platform.
    
    OPTIMIZED: Uses bulk operations and hash-based change detection like stock_update.
    
    Required CSV headers: item_code, units
    Optional headers (at least one required): mrp, cost, stock
    
    Update Types:
    - 'mrp_only': Update MRP and calculate selling_price
    - 'cost_only': Update Cost and calculate converted_cost
    - 'stock_only': Update Stock with WDF conversion for wrap=9900
    - 'mrp_cost': Update both MRP and Cost
    - 'all': Update MRP, Cost, and Stock together
    
    Features:
    - Platform isolation (Pasons vs Talabat)
    - Outlet-platform validation
    - Bulk operations for performance
    - Proper price conversion (WDF for wrap=9900, margin for Talabat)
    - Cascade logic for wrap=9900 parent→children
    """
    from .models import Outlet, Item, ItemOutlet, UploadHistory
    from .utils import calculate_item_selling_price, calculate_item_converted_cost
    from django.contrib import messages
    from django.db import transaction
    from decimal import Decimal, InvalidOperation
    
    if request.method == 'GET':
        outlets = Outlet.objects.filter(is_active=True).order_by('name')
        context = {
            'page_title': 'Product Update',
            'active_nav': 'bulk_operations',
            'outlets': outlets
        }
        return render(request, 'product_update.html', context)
    
    elif request.method == 'POST':
        platform = request.POST.get('platform')
        outlet_id = request.POST.get('outlet')
        update_type = request.POST.get('update_type', 'all')  # mrp_only, cost_only, stock_only, mrp_cost, all
        csv_file = request.FILES.get('csv_file')
        
        # Validate required fields
        if not platform or platform not in ('pasons', 'talabat'):
            messages.error(request, "Please select a valid platform (Pasons or Talabat).")
            return redirect('integration:product_update')
        
        if not outlet_id:
            messages.error(request, "Please select an outlet.")
            return redirect('integration:product_update')
        
        if not csv_file:
            messages.error(request, "Please select a CSV file.")
            return redirect('integration:product_update')
        
        try:
            outlet = Outlet.objects.get(id=outlet_id)
            
            # CRITICAL: Validate outlet matches platform
            if outlet.platforms != platform:
                messages.error(request, f"Outlet '{outlet.name}' does not support {platform.title()} platform.")
                return redirect('integration:product_update')
            
            import csv
            import io
            from django.db.models import Q
            
            csv_content, _encoding_used = decode_csv_upload(csv_file)
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            
            if not csv_reader.fieldnames:
                messages.error(request, "CSV file has no headers")
                return redirect('integration:product_update')
            
            # Import normalize_csv_header for proper BOM/invisible char handling
            from .utils import normalize_csv_header
            headers = [normalize_csv_header(h) for h in csv_reader.fieldnames if h and h.strip()]
            
            # Determine allowed headers based on update_type
            base_headers = {'item_code', 'units'}
            if update_type == 'mrp_only':
                allowed_headers = base_headers | {'mrp'}
                value_headers = {'mrp'}
            elif update_type == 'cost_only':
                allowed_headers = base_headers | {'cost'}
                value_headers = {'cost'}
            elif update_type == 'stock_only':
                allowed_headers = base_headers | {'stock'}
                value_headers = {'stock'}
            elif update_type == 'mrp_cost':
                allowed_headers = base_headers | {'mrp', 'cost'}
                value_headers = {'mrp', 'cost'}
            elif update_type == 'mrp_stock':
                allowed_headers = base_headers | {'mrp', 'stock'}
                value_headers = {'mrp', 'stock'}
            elif update_type == 'cost_stock':
                allowed_headers = base_headers | {'cost', 'stock'}
                value_headers = {'cost', 'stock'}
            else:  # 'all' or default
                allowed_headers = base_headers | {'mrp', 'cost', 'stock'}
                value_headers = {'mrp', 'cost', 'stock'}
            
            # Check for missing required headers
            missing = base_headers - set(headers)
            if missing:
                messages.error(request, f"Missing required columns: {', '.join(sorted(missing))}")
                return redirect('integration:product_update')
            
            # Check that at least one value header is present
            present_value_headers = value_headers & set(headers)
            if not present_value_headers:
                messages.error(request, f"CSV must contain at least one of: {', '.join(sorted(value_headers))}")
                return redirect('integration:product_update')
            
            # Note: Extra columns in CSV are ignored (not an error)
            # This allows users to use the same CSV file with different update types
            
            # Parse all rows
            csv_rows = []
            for row_num, original_row in enumerate(csv_reader, start=2):
                row = {normalize_csv_header(k): v.strip() if v else '' for k, v in original_row.items()}
                if row.get('item_code') and row.get('units'):
                    csv_rows.append((row_num, row))
            
            if not csv_rows:
                messages.warning(request, "No valid rows found in CSV")
                return redirect('integration:product_update')
            
            # Build lookup keys for bulk prefetch - use set directly
            lookup_keys_set = {(r['item_code'], r['units']) for _, r in csv_rows}
            lookup_keys = list(lookup_keys_set)
            
            # OPTIMIZATION: Use .only() to fetch only needed fields
            CHUNK_SIZE = 2000  # Increased for MySQL network latency optimization
            items_dict = {}  # (item_code, units) -> [list of items]
            
            # Determine which item fields we need based on update type
            item_fields = ['id', 'item_code', 'units', 'platform', 'wrap', 'weight_division_factor']
            if 'mrp' in present_value_headers:
                item_fields.extend(['mrp', 'selling_price', 'talabat_margin'])
            if 'cost' in present_value_headers:
                item_fields.extend(['cost', 'converted_cost'])
            if 'stock' in present_value_headers:
                item_fields.append('stock')
            
            for i in range(0, len(lookup_keys), CHUNK_SIZE):
                chunk = lookup_keys[i:i + CHUNK_SIZE]
                
                # SQLite fix: Use separate queries for each (item_code, units) to avoid deep expression trees
                chunk_items = []
                for item_code, units in chunk:
                    items = Item.objects.filter(
                        platform=platform,
                        item_code=item_code,
                        units=units
                    ).only(*item_fields)
                    chunk_items.extend(items)
                for item in chunk_items:
                    key = (item.item_code, item.units)
                    if key not in items_dict:
                        items_dict[key] = []
                    items_dict[key].append(item)
            
            # Pre-load ItemOutlet records with .only()
            all_item_ids = [item.id for items_list in items_dict.values() for item in items_list]
            outlets_map = {}
            
            # Determine which outlet fields we need
            outlet_fields = ['id', 'item_id', 'outlet_id']
            if 'mrp' in present_value_headers:
                outlet_fields.extend(['outlet_mrp', 'outlet_selling_price'])
            if 'cost' in present_value_headers:
                outlet_fields.append('outlet_cost')
            if 'stock' in present_value_headers:
                outlet_fields.append('outlet_stock')
            
            for i in range(0, len(all_item_ids), CHUNK_SIZE):
                chunk_ids = all_item_ids[i:i + CHUNK_SIZE]
                chunk_outlets = ItemOutlet.objects.filter(
                    item_id__in=chunk_ids,
                    outlet=outlet
                ).only(*outlet_fields)
                for io in chunk_outlets:
                    outlets_map[io.item_id] = io
            
            # Collect updates for bulk operations - use SETS for O(1) lookups
            # NOTE: We only update ItemOutlet, not shared Item model (to prevent cross-outlet contamination)
            outlets_to_update_set = set()
            outlets_to_update = []
            outlets_to_create = []
            
            updated_count = 0
            not_found_items = []
            errors = []
            no_change_count = 0
            
            with transaction.atomic():
                for row_num, row in csv_rows:
                    try:
                        item_code = row['item_code']
                        units = row['units']
                        
                        items_list = items_dict.get((item_code, units))
                        if not items_list:
                            not_found_items.append(f"{item_code} ({units})")
                            continue
                        
                        # Process ALL items for this (item_code, units) key
                        for item in items_list:
                            try:
                                item_outlet = outlets_map.get(item.id)
                                is_new_outlet = False
                                
                                if not item_outlet:
                                    # Create new ItemOutlet with Item defaults
                                    item_outlet = ItemOutlet(
                                        item=item,
                                        outlet=outlet,
                                        outlet_stock=item.stock or 0,
                                        outlet_selling_price=item.selling_price or Decimal('0'),
                                        outlet_mrp=item.mrp or Decimal('0'),
                                        outlet_cost=item.cost or Decimal('0'),
                                        is_active_in_outlet=True
                                    )
                                    outlets_map[item.id] = item_outlet
                                    outlets_to_create.append(item_outlet)
                                    is_new_outlet = True
                                
                                outlet_changed = False
                                
                                # CHECK LOCKS before processing updates
                                # CLS locks (item-level)
                                cls_price_locked = bool(getattr(item, 'price_locked', False))
                                cls_status_locked = bool(getattr(item, 'status_locked', False))
                                # BLS locks (outlet-level)
                                bls_price_locked = bool(getattr(item_outlet, 'price_locked', False))
                                bls_status_locked = bool(getattr(item_outlet, 'status_locked', False))
                                
                                # Process MRP - OUTLET-SPECIFIC ONLY (do NOT update shared Item model)
                                # SKIP if price is locked (CLS or BLS)
                                if 'mrp' in present_value_headers:
                                    mrp_str = row.get('mrp', '').strip()
                                    if mrp_str:
                                        # CHECK PRICE LOCK - skip if locked
                                        if cls_price_locked or bls_price_locked:
                                            # Price is locked - skip this update silently
                                            pass
                                        else:
                                            try:
                                                new_mrp = Decimal(mrp_str.replace(',', ''))
                                                if new_mrp < 0:
                                                    new_mrp = Decimal('0')
                                                
                                                mrp_rounded = new_mrp.quantize(Decimal('0.01'))
                                                
                                                # Calculate selling price with proper conversion
                                                new_selling_price = calculate_item_selling_price(item, mrp_rounded, platform)
                                                
                                                # ONLY update outlet MRP and selling price (NOT shared Item model)
                                                current_outlet_mrp = item_outlet.outlet_mrp or Decimal('0')
                                                if current_outlet_mrp != mrp_rounded:
                                                    item_outlet.outlet_mrp = mrp_rounded
                                                    outlet_changed = True
                                                
                                                current_outlet_sp = item_outlet.outlet_selling_price or Decimal('0')
                                                if current_outlet_sp != new_selling_price:
                                                    item_outlet.outlet_selling_price = new_selling_price
                                                    outlet_changed = True
                                            
                                            except InvalidOperation:
                                                errors.append(f"Row {row_num}: Invalid MRP '{mrp_str}'")
                                
                                # Process Cost - OUTLET-SPECIFIC ONLY (do NOT update shared Item model)
                                if 'cost' in present_value_headers:
                                    cost_str = row.get('cost', '').strip()
                                    if cost_str:
                                        try:
                                            new_cost = Decimal(cost_str.replace(',', ''))
                                            if new_cost < 0:
                                                new_cost = Decimal('0')
                                            
                                            # ONLY update outlet cost (NOT shared Item model)
                                            current_outlet_cost = item_outlet.outlet_cost or Decimal('0')
                                            if current_outlet_cost != new_cost:
                                                item_outlet.outlet_cost = new_cost
                                                outlet_changed = True
                                        
                                        except InvalidOperation:
                                            errors.append(f"Row {row_num}: Invalid cost '{cost_str}'")
                                
                                # Process Stock - OUTLET-SPECIFIC ONLY (do NOT update shared Item model)
                                # NOTE: Stock quantity updates even if status_locked (stock number can change)
                                # BUT: is_active_in_outlet stays FALSE if status_locked (item stays Disabled)
                                if 'stock' in present_value_headers:
                                    stock_str = row.get('stock', '').strip()
                                    if stock_str:
                                        try:
                                            csv_stock = int(float(stock_str.replace(',', '')))
                                            if csv_stock < 0:
                                                csv_stock = 0
                                            
                                            # Apply stock conversion based on wrap type
                                            if item.wrap == '9900':
                                                # wrap=9900: stock × WDF (e.g., 10 KG × 4 = 40 packs of 250g)
                                                wdf = item.weight_division_factor or Decimal('1')
                                                new_stock = int(csv_stock * float(wdf))
                                            elif item.wrap == '10000':
                                                # wrap=10000: stock ÷ OCQ (e.g., 50 ÷ 4 = 12.5 cases)
                                                ocq = item.outer_case_quantity or 1
                                                if ocq > 0:
                                                    new_stock = int(csv_stock / ocq)
                                                else:
                                                    new_stock = csv_stock
                                            else:
                                                new_stock = csv_stock
                                            
                                            # ONLY update outlet stock (NOT shared Item model)
                                            if item_outlet.outlet_stock != new_stock:
                                                item_outlet.outlet_stock = new_stock
                                                outlet_changed = True
                                            
                                            # ENFORCE STATUS LOCK: If locked, keep is_active_in_outlet = FALSE
                                            # This prevents stock update from enabling a locked item
                                            if cls_status_locked or bls_status_locked:
                                                # Status is locked - ensure item stays disabled
                                                if item_outlet.is_active_in_outlet:
                                                    item_outlet.is_active_in_outlet = False
                                                    outlet_changed = True
                                        
                                        except ValueError:
                                            errors.append(f"Row {row_num}: Invalid stock '{stock_str}'")
                                
                                # Track changes - O(1) set lookup
                                # NOTE: We no longer update shared Item model, only ItemOutlet
                                if outlet_changed and not is_new_outlet and id(item_outlet) not in outlets_to_update_set:
                                    outlets_to_update_set.add(id(item_outlet))
                                    outlets_to_update.append(item_outlet)
                                
                                if outlet_changed:
                                    updated_count += 1
                                else:
                                    no_change_count += 1
                            
                            except Exception as e:
                                errors.append(f"Row {row_num}: {str(e)}")
                    
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                
                # Bulk operations - OUTLET-SPECIFIC ONLY (no Item model updates)
                if outlets_to_create:
                    ItemOutlet.objects.bulk_create(outlets_to_create, ignore_conflicts=True)
                
                # NOTE: We no longer update shared Item model to prevent cross-outlet contamination
                # All updates are outlet-specific via ItemOutlet
                
                if outlets_to_update:
                    # Only update outlet fields based on what was in the CSV
                    outlet_update_fields = []
                    if 'mrp' in present_value_headers:
                        outlet_update_fields.extend(['outlet_mrp', 'outlet_selling_price'])
                    if 'cost' in present_value_headers:
                        outlet_update_fields.append('outlet_cost')
                    if 'stock' in present_value_headers:
                        outlet_update_fields.append('outlet_stock')
                        # Also update is_active_in_outlet for status lock enforcement
                        outlet_update_fields.append('is_active_in_outlet')
                    
                    if outlet_update_fields:
                        ItemOutlet.objects.bulk_update(outlets_to_update, outlet_update_fields, batch_size=2000)
            
            # Success messages
            if updated_count > 0:
                messages.success(request, f"Updated {updated_count} products at {outlet.name} ({platform.title()}).")
            if no_change_count > 0:
                messages.info(request, f"{no_change_count} products already up-to-date.")
            if not_found_items:
                if len(not_found_items) <= 5:
                    messages.warning(request, f"Items not found: {', '.join(not_found_items)}")
                else:
                    messages.warning(request, f"{len(not_found_items)} items not found.")
            if errors:
                for error in errors[:3]:
                    messages.error(request, error)
                if len(errors) > 3:
                    messages.warning(request, f"And {len(errors) - 3} more errors...")
            
            # Log upload history
            UploadHistory.objects.create(
                file_name=csv_file.name,
                platform=platform,
                outlet=outlet,
                update_type='product',
                records_total=len(csv_rows),
                records_success=updated_count,
                records_failed=len(errors),
                records_skipped=len(not_found_items) + no_change_count,
                status='success' if not errors else ('partial' if updated_count > 0 else 'failed'),
                uploaded_by=request.user if request.user.is_authenticated else None,
            )
            
            return redirect('integration:product_update')
            
        except Outlet.DoesNotExist:
            messages.error(request, "Selected outlet not found.")
        except Exception as e:
            logger.error(f"Product update error: {str(e)}", exc_info=True)
            messages.error(request, f"Error processing CSV: {str(e)}")
        
        return redirect('integration:product_update')
    
    # Fallback for other methods
    outlets = Outlet.objects.filter(is_active=True).order_by('name')
    context = {
        'page_title': 'Product Update',
        'active_nav': 'bulk_operations',
        'outlets': outlets
    }
    return render(request, 'product_update.html', context)


def rules_update_price(request):
    """
    OPTIMIZED Talabat Margin Update via CSV
    Fast bulk update - just updates margin values directly
    """
    from .models import Item
    from django.contrib import messages
    from decimal import Decimal, InvalidOperation
    import csv
    import io
    
    if request.method == 'POST':
        platform = request.POST.get('platform')
        csv_file = request.FILES.get('csv_file')
        
        if platform != 'talabat':
            messages.error(request, "This endpoint is ONLY for Talabat margin updates.")
            return redirect('integration:rules_update_price')
        
        if platform and csv_file:
            try:
                csv_content, _ = decode_csv_upload(csv_file)
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                
                if not csv_reader.fieldnames:
                    messages.error(request, "CSV file has no headers")
                    return redirect('integration:rules_update_price')
                
                from .utils import normalize_csv_header
                headers = [normalize_csv_header(h) for h in csv_reader.fieldnames if h and h.strip()]
                required_headers = {'item_code', 'units', 'sku', 'margin'}
                
                missing = required_headers - set(headers)
                if missing:
                    messages.error(request, f"Missing columns: {', '.join(sorted(missing))}")
                    return redirect('integration:rules_update_price')
                
                # Parse CSV rows (fast - no DB)
                csv_rows = []
                errors = []
                for row_num, original_row in enumerate(csv_reader, start=2):
                    row = {k.strip().lower(): v.strip() if v else '' for k, v in original_row.items()}
                    item_code = row.get('item_code', '')
                    units = row.get('units', '')
                    sku = row.get('sku', '')
                    margin_str = row.get('margin', '')
                    
                    if not item_code or not units or not sku or not margin_str:
                        errors.append(f"Row {row_num}: Missing required field")
                        continue
                    
                    try:
                        margin = Decimal(margin_str)
                        if margin < 0 or margin > 100:
                            errors.append(f"Row {row_num}: Invalid margin")
                            continue
                    except (InvalidOperation, ValueError):
                        errors.append(f"Row {row_num}: Invalid margin")
                        continue
                    
                    csv_rows.append({
                        'item_code': item_code,
                        'units': units,
                        'sku': sku,
                        'margin': margin
                    })
                
                if not csv_rows:
                    messages.warning(request, "No valid rows in CSV")
                    return redirect('integration:rules_update_price')
                
                # BULK FETCH: Single query with .only() for speed
                items_qs = Item.objects.filter(platform='talabat').only(
                    'id', 'item_code', 'units', 'sku', 'talabat_margin'
                )
                
                # Build lookup dict
                items_dict = {(i.item_code, i.units, i.sku): i for i in items_qs}
                
                # Process - FAST: just update margin directly
                items_to_update = []
                updated_count = 0
                not_found_count = 0
                
                for row_data in csv_rows:
                    key = (row_data['item_code'], row_data['units'], row_data['sku'])
                    item = items_dict.get(key)
                    
                    if not item:
                        not_found_count += 1
                        continue
                    
                    # Direct update - no change detection needed
                    item.talabat_margin = row_data['margin']
                    items_to_update.append(item)
                    updated_count += 1
                
                # Bulk update
                if items_to_update:
                    Item.objects.bulk_update(items_to_update, ['talabat_margin'])
                
                # Messages
                if updated_count > 0:
                    messages.success(request, f"Updated {updated_count} Talabat margin(s)")
                if not_found_count > 0:
                    messages.warning(request, f"{not_found_count} item(s) not found")
                if errors:
                    messages.error(request, f"{len(errors)} error(s)")
                
                # Log history
                from .models import UploadHistory
                UploadHistory.objects.create(
                    file_name=csv_file.name,
                    platform='talabat',
                    outlet=None,
                    update_type='rules_price',
                    records_total=len(csv_rows),
                    records_success=updated_count,
                    records_failed=len(errors),
                    records_skipped=not_found_count,
                    status='success' if not errors else 'partial',
                    uploaded_by=request.user if request.user.is_authenticated else None,
                )
                
                return redirect('integration:rules_update_price')
                
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        else:
            messages.error(request, "Please select Talabat and upload CSV.")
    
    context = {
        'page_title': 'Talabat Margin Update',
        'active_nav': 'bulk_operations'
    }
    return render(request, 'rules_update_price.html', context)


@login_required
def rules_update_stock_preview(request):
    """
    OPTIMIZED Preview endpoint for stock conversion rules update
    Uses bulk fetch instead of per-row queries for fast performance
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request'})
    
    platform = request.POST.get('platform')
    csv_file = request.FILES.get('csv_file')
    
    if not platform or not csv_file:
        return JsonResponse({'success': False, 'message': 'Platform and CSV file required'})
    
    try:
        from .models import Item
        from decimal import Decimal, InvalidOperation
        import csv
        import io
        
        csv_content, _encoding_used = decode_csv_upload(csv_file)
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        
        if not csv_reader.fieldnames:
            return JsonResponse({'success': False, 'message': 'CSV has no headers'})
        
        from .utils import normalize_csv_header
        headers = [normalize_csv_header(h) for h in csv_reader.fieldnames if h and h.strip()]
        
        # Header validation
        allowed_headers = {'item_code', 'units', 'sku', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty'}
        required_headers = {'item_code', 'units', 'sku'}
        
        missing_headers = required_headers - set(headers)
        if missing_headers:
            return JsonResponse({'success': False, 'message': f"Missing columns: {', '.join(sorted(missing_headers))}"})
        
        # Parse all CSV rows first (fast - no DB queries)
        csv_rows = []
        errors = []
        for row_num, original_row in enumerate(csv_reader, start=2):
            row = {k.strip().lower(): v.strip() if v else '' for k, v in original_row.items()}
            item_code = row.get('item_code', '')
            units = row.get('units', '')
            sku = row.get('sku', '')
            
            if not item_code or not units or not sku:
                errors.append(f"Row {row_num}: Missing item_code, units, or sku")
                continue
            
            csv_rows.append({
                'row_num': row_num,
                'item_code': item_code,
                'units': units,
                'sku': sku,
                'wdf': row.get('weight_division_factor', ''),
                'ocq': row.get('outer_case_quantity', ''),
                'minqty': row.get('minimum_qty', '')
            })
        
        total_rows = len(csv_rows)
        if total_rows == 0:
            return JsonResponse({'success': True, 'platform': platform, 'total_rows': 0, 'items_with_changes': [], 'total_changes': 0, 'errors': errors})
        
        # BULK FETCH: Single query to get all items
        items_qs = Item.objects.filter(platform=platform).only(
            'id', 'item_code', 'units', 'sku', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty'
        )
        
        # Build lookup dict (item_code, units, sku) -> item
        items_dict = {(i.item_code, i.units, i.sku): i for i in items_qs}
        
        # Compare CSV vs DB (fast - in memory)
        items_with_changes = []
        for row_data in csv_rows:
            key = (row_data['item_code'], row_data['units'], row_data['sku'])
            item = items_dict.get(key)
            
            if not item:
                continue  # Skip not found items in preview
            
            changes = {}
            
            # Check WDF
            if row_data['wdf']:
                try:
                    new_wdf = Decimal(row_data['wdf'])
                    if new_wdf != item.weight_division_factor:
                        changes['wdf'] = {
                            'old': float(item.weight_division_factor) if item.weight_division_factor else None,
                            'new': float(new_wdf)
                        }
                except (InvalidOperation, ValueError):
                    errors.append(f"Row {row_data['row_num']}: Invalid WDF")
                    continue
            
            # Check OCQ
            if row_data['ocq']:
                try:
                    new_ocq = int(row_data['ocq'])
                    if new_ocq != item.outer_case_quantity:
                        changes['ocq'] = {'old': item.outer_case_quantity, 'new': new_ocq}
                except ValueError:
                    errors.append(f"Row {row_data['row_num']}: Invalid OCQ")
                    continue
            
            # Check MinQty
            if row_data['minqty']:
                try:
                    new_minqty = int(row_data['minqty'])
                    if new_minqty != item.minimum_qty:
                        changes['minqty'] = {'old': item.minimum_qty, 'new': new_minqty}
                except ValueError:
                    errors.append(f"Row {row_data['row_num']}: Invalid MinQty")
                    continue
            
            if changes:
                items_with_changes.append({
                    'row': row_data['row_num'],
                    'item_code': row_data['item_code'],
                    'units': row_data['units'],
                    'sku': row_data['sku'],
                    'changes': changes
                })
        
        return JsonResponse({
            'success': True,
            'platform': platform,
            'total_rows': total_rows,
            'items_with_changes': items_with_changes[:100],
            'total_changes': len(items_with_changes),
            'errors': errors[:10] if errors else []
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
def rules_update_stock(request):
    """
    OPTIMIZED Stock conversion rules update - updates weight_division_factor, outer_case_quantity, minimum_qty
    Uses bulk fetch and bulk_update for fast performance. No change detection - just update values directly.
    """
    if request.method == 'POST':
        platform = request.POST.get('platform')
        csv_file = request.FILES.get('csv_file')
        
        if platform and csv_file:
            try:
                from .models import Item
                from django.contrib import messages
                from decimal import Decimal, InvalidOperation
                import csv
                import io
                
                csv_content, _encoding_used = decode_csv_upload(csv_file)
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                
                if not csv_reader.fieldnames:
                    messages.error(request, "CSV file has no headers")
                    return redirect('integration:rules_update_stock')
                
                from .utils import normalize_csv_header
                headers = [normalize_csv_header(h) for h in csv_reader.fieldnames if h and h.strip()]
                
                # Header validation
                allowed_headers = {'item_code', 'units', 'sku', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty'}
                required_headers = {'item_code', 'units', 'sku'}
                
                missing_headers = required_headers - set(headers)
                if missing_headers:
                    messages.error(request, f"Missing required columns: {', '.join(sorted(missing_headers))}")
                    return redirect('integration:rules_update_stock')
                
                # Parse all CSV rows (fast - no DB)
                csv_rows = []
                errors = []
                for row_num, original_row in enumerate(csv_reader, start=2):
                    row = {k.strip().lower(): v.strip() if v else '' for k, v in original_row.items()}
                    item_code = row.get('item_code', '')
                    units = row.get('units', '')
                    sku = row.get('sku', '')
                    
                    if not item_code or not units or not sku:
                        errors.append(f"Row {row_num}: Missing required fields")
                        continue
                    
                    csv_rows.append({
                        'row_num': row_num,
                        'item_code': item_code,
                        'units': units,
                        'sku': sku,
                        'wdf': row.get('weight_division_factor', ''),
                        'ocq': row.get('outer_case_quantity', ''),
                        'minqty': row.get('minimum_qty', '')
                    })
                
                if not csv_rows:
                    messages.warning(request, "No valid rows found in CSV")
                    return redirect('integration:rules_update_stock')
                
                # BULK FETCH: Single query with .only() for speed
                items_qs = Item.objects.filter(platform=platform).only(
                    'id', 'item_code', 'units', 'sku', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty'
                )
                
                # Build lookup dict
                items_dict = {(i.item_code, i.units, i.sku): i for i in items_qs}
                
                # Process rows - FAST: just update values, no change detection needed
                items_to_update = []
                updated_count = 0
                not_found_items = []
                
                for row_data in csv_rows:
                    key = (row_data['item_code'], row_data['units'], row_data['sku'])
                    item = items_dict.get(key)
                    
                    if not item:
                        not_found_items.append(f"{row_data['item_code']} ({row_data['units']})")
                        continue
                    
                    item_changed = False
                    
                    # Update WDF if provided
                    if row_data['wdf']:
                        try:
                            item.weight_division_factor = Decimal(row_data['wdf'])
                            item_changed = True
                        except (InvalidOperation, ValueError):
                            errors.append(f"Row {row_data['row_num']}: Invalid WDF")
                            continue
                    
                    # Update OCQ if provided
                    if row_data['ocq']:
                        try:
                            item.outer_case_quantity = int(row_data['ocq'])
                            item_changed = True
                        except ValueError:
                            errors.append(f"Row {row_data['row_num']}: Invalid OCQ")
                            continue
                    
                    # Update MinQty if provided
                    if row_data['minqty']:
                        try:
                            item.minimum_qty = int(row_data['minqty'])
                            item_changed = True
                        except ValueError:
                            errors.append(f"Row {row_data['row_num']}: Invalid MinQty")
                            continue
                    
                    if item_changed:
                        items_to_update.append(item)
                        updated_count += 1
                
                # Bulk update all items at once
                if items_to_update:
                    Item.objects.bulk_update(
                        items_to_update,
                        ['weight_division_factor', 'outer_case_quantity', 'minimum_qty']
                    )
                
                # Display consolidated messages
                if updated_count > 0:
                    messages.success(request, f"Successfully updated {updated_count} item(s) for {platform.title()} platform")
                
                if not_found_items:
                    messages.warning(request, f"{len(not_found_items)} item(s) not found")
                
                if errors:
                    messages.error(request, f"{len(errors)} error(s) occurred")
                
                # Log upload history
                from .models import UploadHistory
                total_records = updated_count + len(not_found_items) + len(errors)
                upload_status = 'success' if not errors else ('partial' if updated_count else 'failed')
                UploadHistory.objects.create(
                    file_name=csv_file.name,
                    platform=platform,
                    outlet=None,  # Rules update is global
                    update_type='rules_stock',
                    records_total=total_records,
                    records_success=updated_count,
                    records_failed=len(errors),
                    records_skipped=len(not_found_items),
                    status=upload_status,
                    uploaded_by=request.user if request.user.is_authenticated else None,
                )
                
                return redirect('integration:rules_update_stock')
                
            except Exception as e:
                from django.contrib import messages
                messages.error(request, f"Error processing CSV file: {str(e)}")
        else:
            from django.contrib import messages
            if not platform:
                messages.error(request, "Please select a platform.")
            elif not csv_file:
                messages.error(request, "Please select a CSV file.")
    
    context = {
        'page_title': 'Stock Conversion Rules Update',
        'active_nav': 'bulk_operations'
    }
    return render(request, 'rules_update_stock.html', context)



@login_required
def search_product_api(request):
    """
    API endpoint to search for products - supports both single search and bulk loading
    """
    from .models import Item
    from django.db import models
    from django.core.paginator import Paginator
    
    # Determine platform and whether this is bulk mode (item deletion page)
    platform = request.GET.get('platform', '').strip()
    include_inactive = request.GET.get('include_inactive', '').strip() in ('1', 'true', 'True')
    # Bulk mode triggers: explicit pagination or filter parameters
    bulk_mode = (
        ('page' in request.GET) or
        ('page_size' in request.GET) or
        ('item_code' in request.GET) or
        ('description' in request.GET) or
        ('barcode' in request.GET) or
        ('sku' in request.GET) or
        ('price_min' in request.GET) or
        ('price_max' in request.GET) or
        ('stock_min' in request.GET) or
        ('stock_max' in request.GET)
    ) and bool(platform)

    if bulk_mode:
        # Bulk loading for item deletion page with server-side pagination and filters
        try:
            # Page and page size (default 100 per page)
            page = int(request.GET.get('page', '1'))
            page_size = int(request.GET.get('page_size', '100'))
            if page_size <= 0:
                page_size = 100
            if page_size > 1000:  # prevent excessive queries
                page_size = 1000

            # Filters
            item_code = request.GET.get('item_code', '').strip()
            description = request.GET.get('description', '').strip()
            barcode = request.GET.get('barcode', '').strip()
            sku = request.GET.get('sku', '').strip()
            price_min = request.GET.get('price_min', '').strip()
            price_max = request.GET.get('price_max', '').strip()
            stock_min = request.GET.get('stock_min', '').strip()
            stock_max = request.GET.get('stock_max', '').strip()

            qs = Item.objects.all() if include_inactive else Item.objects.filter(is_active=True)
            if platform in ('pasons', 'talabat'):
                qs = qs.filter(platform=platform)
            elif platform == 'all':
                qs = qs.filter(platform__in=['pasons', 'talabat'])

            # Apply filters
            if item_code:
                qs = qs.filter(item_code__icontains=item_code)
            if description:
                qs = qs.filter(description__icontains=description)
            if barcode:
                qs = qs.filter(barcode__icontains=barcode)
            if sku:
                qs = qs.filter(sku__icontains=sku)
            # Numeric filters
            try:
                if price_min:
                    qs = qs.filter(selling_price__gte=float(price_min))
            except ValueError:
                pass
            try:
                if price_max:
                    qs = qs.filter(selling_price__lte=float(price_max))
            except ValueError:
                pass
            try:
                if stock_min:
                    qs = qs.filter(stock__gte=int(stock_min))
            except ValueError:
                pass
            try:
                if stock_max:
                    qs = qs.filter(stock__lte=int(stock_max))
            except ValueError:
                pass

            qs = qs.order_by('item_code')

            # Pagination
            paginator = Paginator(qs, page_size)
            page_obj = paginator.get_page(page)

            items_data = []
            for item in page_obj.object_list:
                combination_key = f"{item.item_code}|{item.description}|{item.sku}"
                items_data.append({
                    'id': item.id,
                    'item_code': item.item_code,
                    'description': item.description,
                    'pack_description': item.pack_description or '',
                    'sku': item.sku,
                    'units': item.units,
                    'selling_price': float(item.selling_price),
                    'stock': item.stock,
                    'mrp': float(item.mrp),
                    'cost': float(item.cost),
                    'is_active': item.is_active,
                    # CLS flags for UI consistency
                    'price_locked': bool(getattr(item, 'price_locked', False)),
                    'status_locked': bool(getattr(item, 'status_locked', False)),
                    'barcode': item.barcode or '',
                    'wrap': item.wrap or '',
                    'weight_division_factor': item.weight_division_factor,
                    'outer_case_quantity': item.outer_case_quantity,
                    'minimum_qty': item.minimum_qty,
                    'talabat_margin': float(item.effective_talabat_margin) if platform == 'talabat' and item.effective_talabat_margin is not None else None,
                    'combination_key': combination_key
                })

            return JsonResponse({
                'success': True,
                'items': items_data,
                'page': page_obj.number,
                'page_size': page_size,
                'total_items': paginator.count,
                'total_pages': paginator.num_pages,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error loading items: {str(e)}'
            })
    
    # Single product search (original functionality)
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'found': False, 'message': 'No search query provided'})
    
    try:
        # Base queryset restricted by platform. Honor include_inactive for consistency.
        qs = Item.objects.all() if include_inactive else Item.objects.filter(is_active=True)
        if platform in ('pasons', 'talabat'):
            qs = qs.filter(platform=platform)
        elif platform == 'all':
            qs = qs.filter(platform__in=['pasons', 'talabat'])
        # Search by item_code, sku, barcode, or description
        qs = qs.filter(
            models.Q(item_code__iexact=query) |
            models.Q(sku__iexact=query) |
            models.Q(barcode__iexact=query) |
            models.Q(description__icontains=query) |
            models.Q(item_code__icontains=query) |
            models.Q(sku__icontains=query)
        )[:20]  # Limit to 20 results
        
        if qs.exists():
            items_data = []
            for item in qs:
                # Create unique combination key for item_code + item_name + sku
                combination_key = f"{item.item_code}|{item.description}|{item.sku}"
                
                items_data.append({
                    'id': item.id,
                    'item_code': item.item_code,
                    'description': item.description,
                    'pack_description': item.pack_description or '',
                    'sku': item.sku,
                    'units': item.units,
                    'selling_price': float(item.selling_price),
                    'stock': item.stock,
                    'mrp': float(item.mrp),
                    'cost': float(item.cost),
                    'price_locked': bool(getattr(item, 'price_locked', False)),
                    'status_locked': bool(getattr(item, 'status_locked', False)),
                    'barcode': item.barcode or '',
                    'wrap': item.wrap or '',
                    'weight_division_factor': item.weight_division_factor,
                    'outer_case_quantity': item.outer_case_quantity,
                    'minimum_qty': item.minimum_qty,
                    'talabat_margin': float(item.effective_talabat_margin) if platform == 'talabat' and item.effective_talabat_margin is not None else None,
                    'combination_key': combination_key
                })
            
            return JsonResponse({
                'success': True,
                'items': items_data,
                'total_count': len(items_data)
            })
        else:
            return JsonResponse({'success': False, 'message': 'No items found matching your search'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Search error: {str(e)}'})


def calculate_outlet_enabled_status(item, outlet_stock):
    """
    Calculate if an outlet should show as Enabled or Disabled
    based on stock and minimum_qty.
    
    Rules:
    1. stock ≤ 0 → Disabled (No stock = always disabled)
    2. outlet_stock is ALREADY converted during stock update:
       - wrap=9900: outlet_stock = CSV_stock × WDF (already in packs)
       - wrap=10000: outlet_stock = CSV_stock ÷ OCQ (already in cases)
    3. If outlet_stock > minimum_qty → Enabled
    4. If outlet_stock ≤ minimum_qty → Disabled
    
    Examples (wrap=9900, 250gm item with WDF=4):
    - CSV stock=3 KG → outlet_stock=12 packs, min_qty=10 → 12>10 → Enabled
    - CSV stock=2 KG → outlet_stock=8 packs, min_qty=10 → 8≤10 → Disabled
    
    Examples (wrap=10000, OCQ=100):
    - CSV stock=1800 → outlet_stock=18 cases, min_qty=3 → 18>3 → Enabled
    - CSV stock=200 → outlet_stock=2 cases, min_qty=3 → 2≤3 → Disabled
    
    Returns:
        bool: True = Enabled (stock_status=1), False = Disabled (stock_status=0)
    """
    stock = outlet_stock or 0
    
    # Rule 1: No stock or negative stock = Always Disabled
    if stock <= 0:
        return False
    
    # Rule 2: outlet_stock is ALREADY converted (no further division needed)
    # wrap=9900: already multiplied by WDF during stock update
    # wrap=10000: already divided by OCQ during stock update
    
    # Rule 3: Check if stock is GREATER THAN minimum_qty requirement
    min_qty = item.minimum_qty
    if min_qty is not None and min_qty > 0:
        if stock <= min_qty:  # Must be GREATER than (not equal)
            return False
    
    # All checks passed = Enabled
    return True


@login_required
def item_outlets_api(request):
    """
    Return outlet availability for a given item on a given platform.
    Accepts: platform (required), item_code + units OR item_id
    Responds with: product summary and list of outlets with price/stock.
    
    IMPORTANT: item_code + units together form a unique item identifier!
    Same item_code can have multiple units (e.g., 100010915 with PCS, DZN, OUT, JAR)
    
    NOTE: 'active' field is now AUTO-CALCULATED based on:
    - stock > 0
    - stock >= outer_case_quantity (if set)
    - stock >= minimum_qty (if set)
    """
    from .models import Item, ItemOutlet, Outlet

    platform = request.GET.get('platform', '').strip()
    item_code = request.GET.get('item_code', '').strip()
    units = request.GET.get('units', '').strip()  # NEW: Accept units parameter
    sku = request.GET.get('sku', '').strip()  # NEW: Accept SKU for unique identification (wrap=9900)
    item_id = request.GET.get('item_id', '').strip()
    include_inactive = request.GET.get('include_inactive', '').strip() in ('1', 'true', 'True')

    # STRICT ISOLATION: Only 'pasons' or 'talabat' allowed
    if platform not in ('pasons', 'talabat'):
        return JsonResponse({'success': False, 'message': 'Invalid or missing platform'})

    try:
        item = None
        if item_id:
            try:
                if include_inactive:
                    item = Item.objects.filter(pk=int(item_id)).first()
                else:
                    item = Item.objects.filter(pk=int(item_id), is_active=True).first()
            except ValueError:
                item = None
        if item is None and item_code:
            # FIXED: Filter by item_code, units, sku, AND platform for unique identification
            # Platform is CRITICAL - same item_code can exist on both Pasons and Talabat!
            item_filter = {'item_code__iexact': item_code}
            
            # CRITICAL FIX: Filter by platform to ensure we get the right item!
            # Without this, Talabat dashboard might find a Pasons item and show no outlets!
            if platform in ('pasons', 'talabat'):
                item_filter['platform'] = platform
            
            if sku:  # SKU is the most specific - use it first
                item_filter['sku__iexact'] = sku
            elif units:  # Fallback to units if no SKU
                item_filter['units__iexact'] = units
            if not include_inactive:
                item_filter['is_active'] = True
            item = Item.objects.filter(**item_filter).first()

        if item is None:
            return JsonResponse({'success': True, 'product': None, 'outlets': []})

        # Platform scoping: STRICT platform isolation
        # Pasons and Talabat are completely separate platforms - no overlap!
        pfilter = {'outlet__platforms': platform}  # Strict isolation

        io_qs = ItemOutlet.objects.select_related('outlet').filter(
            item=item,
            outlet__is_active=True,
            **pfilter
        ).order_by('outlet__name')

        outlets = []
        if io_qs.exists():
            for io in io_qs:
                # Display outlet-specific selling price
                # Always show original outlet_selling_price - promo price is separate
                if io.outlet_selling_price is not None:
                    price = io.outlet_selling_price
                else:
                    price = 0.00
                
                # Calculate enabled status based on stock rules
                # Only show as Enabled if: has stock AND meets quantity requirements
                calculated_enabled = calculate_outlet_enabled_status(item, io.outlet_stock)
                
                # Final status: Must pass both manual flag AND stock rules
                # If manually disabled (is_active_in_outlet=False), stay disabled
                # If stock rules fail, show disabled regardless of manual flag
                effective_active = io.is_active_in_outlet and calculated_enabled
                
                outlets.append({
                    'outlet_name': io.outlet.name,
                    'store_id': io.outlet.store_id,
                    'location': io.outlet.location,
                    'platform': io.outlet.platforms,
                    'price': float(price),
                    'stock': io.outlet_stock,
                    'active': effective_active,  # Now auto-calculated!
                    'stock_status_reason': 'ok' if calculated_enabled else 'insufficient_stock',
                    # Cost fields (OUTLET-LEVEL)
                    # wrap=9900: converted_cost = outlet_cost / WDF
                    # wrap=10000: converted_cost = outlet_cost (no conversion)
                    'outlet_cost': float(io.outlet_cost) if io.outlet_cost is not None else 0.00,
                    'outlet_converted_cost': (
                        float((io.outlet_cost / item.weight_division_factor).quantize(Decimal('0.001')))
                        if io.outlet_cost is not None and item.wrap == '9900' and item.weight_division_factor
                        else (float(io.outlet_cost) if io.outlet_cost is not None else 0.00)
                    ),
                    # MRP and S.Price (OUTLET-SPECIFIC, not global)
                    'outlet_mrp': float(io.outlet_mrp) if io.outlet_mrp is not None else 0.00,
                    'outlet_selling_price': float(price),
                    # BLS states
                    'locked': bool(getattr(io, 'price_locked', False)),
                    'price_locked': bool(getattr(io, 'price_locked', False)),
                    'status_locked': bool(getattr(io, 'status_locked', False)),
                    'associated': True,
                })
        # No fallback - if no ItemOutlet records, return empty outlets list
        # Outlets only appear AFTER price/stock is updated via price-update

        product = {
            'item_code': item.item_code,
            'description': item.description,
            'pack_description': item.pack_description or '',
            'units': item.units,  # ADDED: Include units for unique identification
            'sku': item.sku,  # ADDED: SKU for unique item identification
            'wrap': item.wrap,  # ADDED: wrap type (9900 or 10000)
            'mrp': float(item.mrp) if item.mrp is not None else 0.00,
            'cost': float(item.cost) if item.cost is not None else 0.00,
            'converted_cost': float(item.converted_cost) if item.converted_cost is not None else None,
            'selling_price': float(item.selling_price) if item.selling_price is not None else 0.00,
            'weight_division_factor': float(item.weight_division_factor) if item.weight_division_factor is not None else None,
            # Talabat margin (uses effective_talabat_margin which auto-detects if not set)
            'talabat_margin': float(item.effective_talabat_margin) if item.platform == 'talabat' and item.effective_talabat_margin is not None else None,
            # CLS states
            'price_locked': bool(getattr(item, 'price_locked', False)),
            'status_locked': bool(getattr(item, 'status_locked', False)),
        }

        return JsonResponse({'success': True, 'product': product, 'outlets': outlets})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': f'Outlet availability error: {str(e)}'})


@login_required
def outlet_price_update_api(request):
    """
    Update outlet-specific selling price for an item.
    Expects POST with: item_code or item_id, store_id, price (or new_price).
    """
    from .models import Item, Outlet, ItemOutlet

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST method allowed'})

    try:
        item_code = (request.POST.get('item_code') or '').strip()
        item_id = (request.POST.get('item_id') or '').strip()
        units = (request.POST.get('units') or '').strip()  # NEW: Accept units
        store_id = (request.POST.get('store_id') or '').strip()
        platform = (request.POST.get('platform') or '').strip().lower()
        price_str = (request.POST.get('price') or request.POST.get('new_price') or '').strip()

        if not store_id:
            return JsonResponse({'success': False, 'message': 'store_id is required'})
        if not (item_code or item_id):
            return JsonResponse({'success': False, 'message': 'item_code or item_id is required'})
        if not price_str:
            return JsonResponse({'success': False, 'message': 'price is required'})
        if platform not in ('pasons', 'talabat'):
            return JsonResponse({'success': False, 'message': 'Invalid or missing platform parameter'})

        try:
            new_price = Decimal(price_str)
            if new_price < 0:
                return JsonResponse({'success': False, 'message': 'price must be non-negative'})
        except (InvalidOperation, ValueError):
            return JsonResponse({'success': False, 'message': 'Invalid price format'})

        # Resolve item with platform filter
        item = None
        if item_id:
            try:
                item = Item.objects.filter(pk=int(item_id), platform=platform, is_active=True).first()
            except ValueError:
                item = None
        if item is None and item_code:
            # FIXED: Filter by BOTH item_code AND units for unique identification
            item_filter = {'item_code__iexact': item_code, 'platform': platform, 'is_active': True}
            if units:  # If units provided, use it for exact match
                item_filter['units__iexact'] = units
            item = Item.objects.filter(**item_filter).first()
        if item is None:
            return JsonResponse({'success': False, 'message': 'Item not found or inactive'})

        # Resolve outlet
        outlet = Outlet.objects.filter(store_id=store_id, is_active=True).first()
        if outlet is None:
            return JsonResponse({'success': False, 'message': 'Outlet not found or inactive'})

        # CHECK CLS PRICE LOCK FIRST - before any ItemOutlet operations
        if bool(getattr(item, 'price_locked', False)):
            return JsonResponse({'success': False, 'message': 'Price is locked at item level (CLS). Unlock to edit.'})

        # Resolve relation - auto-link if item is already on this platform (via other outlets)
        io = ItemOutlet.objects.filter(item=item, outlet=outlet).first()
        if io is None:
            # Check if item is already associated with this platform via other outlets
            platform = outlet.platforms
            existing_on_platform = ItemOutlet.objects.filter(
                item=item,
                outlet__platforms=platform
            ).exists()
            
            if not existing_on_platform:
                # Item is NOT on this platform yet - don't auto-create (would change platform count)
                return JsonResponse({
                    'success': False, 
                    'message': f'Item is not available on {platform} platform. Please add it via bulk upload first.'
                })
            
            # Item IS on this platform - safe to link to another outlet on same platform
            # Create WITHOUT price first, set price after lock check
            io = ItemOutlet.objects.create(
                item=item,
                outlet=outlet,
                is_active_in_outlet=True,
                outlet_selling_price=item.selling_price  # Default to item price
            )

        # Enforce BLS price lock (CLS already checked above)
        if bool(getattr(io, 'price_locked', False)):
            return JsonResponse({'success': False, 'message': 'Price is locked for this outlet (BLS). Unlock to edit.'})

        io.outlet_selling_price = new_price
        io.save()

        return JsonResponse({
            'success': True,
            'message': 'Outlet price updated',
            'item_code': item.item_code,
            'store_id': outlet.store_id,
            'price': float(new_price)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error updating outlet price: {str(e)}'})


@login_required
def outlet_lock_toggle_api(request):
    """
    Toggle outlet-level status/price lock for an item.
    Current model does not define explicit lock fields; we support status lock
    by toggling `is_active_in_outlet`. Price lock is acknowledged but not persisted.

    Expects POST with: item_code or item_id, store_id, lock_type ('status'|'price'),
    optional 'value' ('true'|'false' or 'lock'|'unlock').
    """
    from .models import Item, Outlet, ItemOutlet

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST method allowed'})

    try:
        item_code = (request.POST.get('item_code') or '').strip()
        item_id = (request.POST.get('item_id') or '').strip()
        units = (request.POST.get('units') or '').strip()  # NEW: Accept units
        store_id = (request.POST.get('store_id') or '').strip()
        lock_type = (request.POST.get('lock_type') or 'status').strip().lower()
        value_raw = (request.POST.get('value') or '').strip().lower()

        if not store_id:
            return JsonResponse({'success': False, 'message': 'store_id is required'})
        if not (item_code or item_id):
            return JsonResponse({'success': False, 'message': 'item_code or item_id is required'})

        # Resolve outlet FIRST to get platform
        outlet = Outlet.objects.filter(store_id=store_id, is_active=True).first()
        if outlet is None:
            return JsonResponse({'success': False, 'message': 'Outlet not found or inactive'})
        
        platform = outlet.platforms

        # Resolve item - FILTER BY PLATFORM to get correct item
        item = None
        if item_id:
            try:
                item = Item.objects.filter(pk=int(item_id), is_active=True, platform=platform).first()
            except ValueError:
                item = None
        if item is None and item_code:
            # FIXED: Filter by item_code, units, AND platform for unique identification
            item_filter = {'item_code__iexact': item_code, 'is_active': True, 'platform': platform}
            if units:  # If units provided, use it for exact match
                item_filter['units__iexact'] = units
            item = Item.objects.filter(**item_filter).first()
        if item is None:
            return JsonResponse({'success': False, 'message': f'Item not found on {platform} platform or inactive'})

        # Resolve relation - auto-link if item is already on this platform (via other outlets)
        io = ItemOutlet.objects.filter(item=item, outlet=outlet).first()
        if io is None:
            # Check if item is already associated with this platform via other outlets
            existing_on_platform = ItemOutlet.objects.filter(
                item=item,
                outlet__platforms=platform
            ).exists()
            
            if not existing_on_platform:
                # Item is NOT on this platform yet - don't auto-create (would change platform count)
                return JsonResponse({
                    'success': False, 
                    'message': f'Item is not available on {platform} platform. Please add it via bulk upload first.'
                })
            
            # Item IS on this platform - safe to link to another outlet on same platform
            io = ItemOutlet.objects.create(
                item=item,
                outlet=outlet,
                is_active_in_outlet=True
            )

        # Parse desired value
        desired = None
        if value_raw in ('true', 'lock', 'locked', '1'):
            desired = True
        elif value_raw in ('false', 'unlock', 'unlocked', '0'):
            desired = False

        # Prevent manual BLS Status changes when CLS Status Lock is enabled
        if lock_type == 'status' and bool(getattr(item, 'status_locked', False)):
            return JsonResponse({
                'success': False,
                'message': 'Central Status Lock is enabled for this item; outlet status cannot be changed.'
            })

        # Prevent manual BLS Price changes when CLS Price Lock is enabled
        if lock_type == 'price' and bool(getattr(item, 'price_locked', False)):
            return JsonResponse({
                'success': False,
                'message': 'Central Price Lock is enabled for this item; outlet price lock cannot be changed.'
            })

        if lock_type == 'status':
            # BLS: toggle status lock
            # CHECKED (locked=True) => Force DISABLED (is_active_in_outlet=False)
            # UNCHECKED (locked=False) => Enable based on stock rules
            current = bool(getattr(io, 'status_locked', False))
            new_val = (not current) if desired is None else bool(desired)
            io.status_locked = new_val
            
            if new_val:
                # CHECKED: Force disable (ignore stock rules)
                io.is_active_in_outlet = False
            else:
                # UNCHECKED: Enable based on stock rules
                calculated_enabled = calculate_outlet_enabled_status(item, io.outlet_stock)
                io.is_active_in_outlet = calculated_enabled
            
            io.save(update_fields=['status_locked', 'is_active_in_outlet'])
            
            # Return effective status for UI update
            effective_active = io.is_active_in_outlet
            
            return JsonResponse({
                'success': True,
                'message': 'Outlet status lock toggled',
                'store_id': outlet.store_id,
                'item_code': item.item_code,
                'active_in_outlet': effective_active,  # Now returns calculated status!
                'status_locked': io.status_locked,
            })
        elif lock_type == 'price':
            # BLS: toggle price lock (no immediate UI badge change)
            current = bool(getattr(io, 'price_locked', False))
            new_val = (not current) if desired is None else bool(desired)
            io.price_locked = new_val
            io.save(update_fields=['price_locked'])
            return JsonResponse({
                'success': True,
                'message': 'Outlet price lock toggled',
                'store_id': outlet.store_id,
                'item_code': item.item_code,
                'price_locked': io.price_locked,
            })
        else:
            return JsonResponse({'success': False, 'message': 'Invalid lock_type; use status or price'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error toggling outlet lock: {str(e)}'})

@login_required
def cls_lock_toggle_api(request):
    """
    Toggle Central Locking System (CLS) locks for an item.
    Supports both Status and Price locks.

    POST fields:
      - item_code or item_id
      - lock_type: 'status' | 'price' (optional; default 'status')
      - value: 'true'|'false'|'lock'|'unlock' (optional; toggles when missing)
      - Alternatively for price: 'price_locked' can be provided ('on'|'true'|'1')
    """
    from .models import Item, ItemOutlet

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST method allowed'})

    try:
        item_code = (request.POST.get('item_code') or '').strip()
        item_id = (request.POST.get('item_id') or '').strip()
        units = (request.POST.get('units') or '').strip()  # For unique item identification
        platform = (request.POST.get('platform') or '').strip()  # Platform filter
        lock_type = (request.POST.get('lock_type') or 'status').strip().lower()
        value_raw = (request.POST.get('value') or '').strip().lower()

        # Helper to parse boolean-ish values
        def _parse_bool(val):
            return str(val).lower() in ('on', 'true', '1', 'yes', 'locked')

        # Resolve item - FILTER BY PLATFORM for correct item
        item = None
        if item_id:
            try:
                item_filter = {'pk': int(item_id), 'is_active': True}
                if platform:
                    item_filter['platform'] = platform
                item = Item.objects.filter(**item_filter).first()
            except ValueError:
                item = None
        if item is None and item_code:
            item_filter = {'item_code__iexact': item_code, 'is_active': True}
            if platform:
                item_filter['platform'] = platform
            if units:
                item_filter['units__iexact'] = units
            item = Item.objects.filter(**item_filter).first()
        if item is None:
            return JsonResponse({'success': False, 'message': 'Item not found or inactive'})

        # Parse desired value
        desired = None
        if value_raw in ('true', 'lock', 'locked', '1'):
            desired = True
        elif value_raw in ('false', 'unlock', 'unlocked', '0'):
            desired = False

        if lock_type == 'status':
            current = bool(getattr(item, 'status_locked', False))
            new_val = (not current) if desired is None else bool(desired)

            # Update item CLS status lock
            item.status_locked = new_val
            item.save(update_fields=['status_locked'])

            # Cascade to all ItemOutlet rows via model helper
            cascade_success = True
            try:
                item.cascade_cls_status_to_outlets(new_val)
            except Exception as e:
                logger.warning(f"CLS status cascade failed for item {item.item_code}: {e}")
                cascade_success = False

            # Verify cascade by checking ItemOutlet records
            outlet_locks = ItemOutlet.objects.filter(
                item=item,
                outlet__platforms=item.platform
            ).values('outlet__name', 'status_locked')
            
            outlet_lock_summary = {
                str(ol['outlet__name']): ol['status_locked']
                for ol in outlet_locks
            }

            return JsonResponse({
                'success': True,
                'message': 'CLS Status Lock updated',
                'item_code': item.item_code,
                'status_locked': item.status_locked,
                'cascade_success': cascade_success,
                'outlet_locks': outlet_lock_summary,  # ← Frontend can refresh with this
            })

        elif lock_type == 'price':
            # Allow either explicit 'price_locked' param or generic 'value'
            if request.POST.get('price_locked') is not None and desired is None:
                desired = _parse_bool(request.POST.get('price_locked'))
            current = bool(getattr(item, 'price_locked', False))
            new_val = (not current) if desired is None else bool(desired)

            # Update item CLS price lock
            item.price_locked = new_val
            item.save(update_fields=['price_locked'])

            # Cascade to outlets' price locks via model helper
            cascade_success = True
            try:
                item.cascade_cls_price_to_outlets(new_val)
            except Exception as e:
                logger.warning(f"CLS price cascade failed for item {item.item_code}: {e}")
                cascade_success = False

            # Verify cascade by checking ItemOutlet records
            outlet_locks = ItemOutlet.objects.filter(
                item=item,
                outlet__platforms=item.platform
            ).values('outlet__name', 'price_locked')
            
            outlet_lock_summary = {
                str(ol['outlet__name']): ol['price_locked']
                for ol in outlet_locks
            }

            return JsonResponse({
                'success': True,
                'message': 'CLS Price Lock updated',
                'item_code': item.item_code,
                'price_locked': item.price_locked,
                'cascade_success': cascade_success,
                'outlet_locks': outlet_lock_summary,  # ← Frontend can refresh with this
            })
        else:
            return JsonResponse({'success': False, 'message': 'Invalid lock_type; use status or price'})

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'CLS lock toggle error: {str(e)}'})

@login_required
def save_product_api(request):
    """
    API endpoint to save/update a product
    Supports both JSON body and form data
    """
    from .models import Item, ItemOutlet
    import json
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST method allowed'})
    
    try:
        # Parse request data - support both JSON and form data
        content_type = request.content_type or ''
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
        else:
            # Fallback to form data
            data = request.POST
        
        # Helper to get value from parsed data
        def get_val(key, default=''):
            val = data.get(key, default)
            if val is None:
                return default
            return val
        
        item_code = str(get_val('item_code', '')).strip()
        if not item_code:
            return JsonResponse({'success': False, 'message': 'Item code is required'})
        
        # Normalize incoming fields and provide sensible fallbacks
        name = str(get_val('description', get_val('name', '')))
        pack_description = str(get_val('pack_description', ''))
        stock_quantity = int(get_val('stock', get_val('stock_quantity', 0)) or 0)
        selling_price = float(get_val('selling_price', 0) or 0)
        mrp = float(get_val('mrp', 0) or 0)
        sku = str(get_val('sku', ''))
        barcode = str(get_val('barcode', ''))
        # Support both 'cost' and 'cost_price' field names
        cost_str = get_val('cost', get_val('cost_price', 0))
        cost = float(cost_str or 0)
        wrap = str(get_val('wrap', ''))
        weight_division_factor_str = str(get_val('weight_division_factor', ''))
        weight_division_factor = float(weight_division_factor_str) if weight_division_factor_str else None
        outer_case_quantity_str = str(get_val('outer_case_quantity', ''))
        outer_case_quantity = int(outer_case_quantity_str) if outer_case_quantity_str else None
        minimum_qty_str = str(get_val('minimum_qty', ''))
        minimum_qty = int(minimum_qty_str) if minimum_qty_str else None
        is_active_val = get_val('is_active', '')
        is_active = str(is_active_val).lower() in ('on', 'true', '1', 'yes') if is_active_val else True
        # CLS toggles
        def _parse_bool(val):
            return str(val).lower() in ('on', 'true', '1', 'yes')
        price_locked_flag = _parse_bool(get_val('price_locked', ''))
        status_locked_flag = _parse_bool(get_val('status_locked', ''))
        # Note: stock_status is auto-calculated on frontend based on outlet stock (not stored in DB)

        # Validate wrap strictly when provided
        if wrap and wrap not in ALLOWED_WRAP_VALUES:
            return JsonResponse({'success': False, 'message': "Wrap must be 9900 or 10000"})
        
        # Get platform from request
        platform = str(get_val('platform', '')).strip()
        if platform not in ('pasons', 'talabat'):
            return JsonResponse({'success': False, 'message': 'Invalid or missing platform parameter'})
        
        # Try to get existing item or create new one - WITH PLATFORM FILTER
        item, created = Item.objects.get_or_create(
            platform=platform,  # ✓ Add platform filter for platform isolation
            item_code=item_code,
            defaults={
                'platform': platform,  # ✓ Set platform in defaults
                'description': name,
                'pack_description': pack_description,
                'stock': stock_quantity,
                'selling_price': selling_price,
                'mrp': mrp,
                'sku': sku,
                'barcode': barcode,
                'cost': cost,
                'wrap': wrap,
                'weight_division_factor': weight_division_factor,
                'outer_case_quantity': outer_case_quantity,
                'minimum_qty': minimum_qty,
                'units': str(get_val('units', '')),
                'is_active': is_active,
                'price_locked': price_locked_flag,
                'status_locked': status_locked_flag,
                # stock_status is frontend-calculated field (not stored)
            }
        )
        
        # Calculate converted_cost = cost / weight_division_factor
        if weight_division_factor and weight_division_factor > 0:
            from decimal import Decimal
            converted_cost = Decimal(str(cost)) / Decimal(str(weight_division_factor))
            item.converted_cost = converted_cost
        else:
            item.converted_cost = None
        
        # Track original CLS status to determine cascade needs
        original_status_locked = item.status_locked if not created else None

        if not created:
            # Update existing item
            item.description = name or item.description
            item.pack_description = pack_description or item.pack_description
            item.stock = stock_quantity if 'stock' in data or 'stock_quantity' in data else item.stock
            item.selling_price = selling_price if 'selling_price' in data else item.selling_price
            item.mrp = mrp if 'mrp' in data else item.mrp
            item.sku = sku or item.sku
            item.barcode = barcode or item.barcode
            item.cost = cost if ('cost' in data or 'cost_price' in data) else item.cost
            item.wrap = wrap or item.wrap
            item.weight_division_factor = weight_division_factor if weight_division_factor_str else item.weight_division_factor
            item.outer_case_quantity = outer_case_quantity if outer_case_quantity_str else item.outer_case_quantity
            item.minimum_qty = minimum_qty if minimum_qty_str else item.minimum_qty
            item.units = str(get_val('units', '')) or item.units
            item.is_active = is_active
            # Only update locks if provided
            if 'price_locked' in data:
                item.price_locked = price_locked_flag
            if 'status_locked' in data:
                item.status_locked = status_locked_flag
            
            # Recalculate converted_cost when weight_division_factor is updated
            if weight_division_factor_str:
                if weight_division_factor and weight_division_factor > 0:
                    from decimal import Decimal
                    item.converted_cost = Decimal(str(item.cost)) / Decimal(str(weight_division_factor))
                else:
                    item.converted_cost = None
            
            item.save()
        else:
            # Item created with initial flags already set in get_or_create defaults
            pass

        # Cascade CLS Status Lock to all ItemOutlet rows for this item when provided
        if 'status_locked' in data:
            try:
                ItemOutlet.objects.filter(
                    item=item,
                    outlet__platforms=platform  # STRICT platform isolation
                ).update(
                    status_locked=status_locked_flag,
                    is_active_in_outlet=(not status_locked_flag)
                )
            except Exception as e:
                # Do not fail save; report cascade issue in message for awareness
                logger.warning(f"CLS status cascade failed for item {item.item_code}: {e}")
        
        action = 'created' if created else 'updated'
        return JsonResponse({
            'success': True, 
            'message': f'Product {action} successfully!',
            'item_code': item.item_code
        })
        
    except ValueError as e:
        return JsonResponse({'success': False, 'message': f'Invalid data format: {str(e)}'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error saving product: {str(e)}'})


@login_required
def item_deletion(request):
    """
    Item deletion page view that renders the item deletion template
    """
    context = {
        'page_title': 'Item Deletion Management',
        'active_page': 'item_deletion'
    }
    return render(request, 'item_deletion.html', context)


@login_required
def shop_integration(request):
    """
    Shop Integration page with tabs for different integration methods
    Tabs: CSV/Manual | SFTP | API
    Sub-tabs: Pasons | Talabat
    """
    from .models import UploadHistory, Outlet, ExportHistory
    
    # Get recent uploads for each platform (load more for pagination)
    pasons_uploads = UploadHistory.objects.filter(platform='pasons').select_related('outlet')[:100]
    talabat_uploads = UploadHistory.objects.filter(platform='talabat').select_related('outlet')[:100]
    global_uploads = UploadHistory.objects.filter(platform='global').select_related('outlet')[:100]
    
    # Get export history for each platform
    pasons_exports = ExportHistory.objects.filter(platform='pasons').select_related('outlet').order_by('-export_timestamp')[:50]
    talabat_exports = ExportHistory.objects.filter(platform='talabat').select_related('outlet').order_by('-export_timestamp')[:50]
    
    # Get outlets for each platform (for export dropdown)
    pasons_outlets = Outlet.objects.filter(platforms='pasons', is_active=True).order_by('name')
    talabat_outlets = Outlet.objects.filter(platforms='talabat', is_active=True).order_by('name')
    
    context = {
        'page_title': 'Shop Integration',
        'active_page': 'shop_integration',
        'pasons_uploads': pasons_uploads,
        'talabat_uploads': talabat_uploads,
        'global_uploads': global_uploads,
        'pasons_exports': pasons_exports,
        'talabat_exports': talabat_exports,
        'pasons_outlets': pasons_outlets,
        'talabat_outlets': talabat_outlets,
    }
    return render(request, 'shop_integration.html', context)


@login_required
def preview_csv_api(request):
    """
    API endpoint to preview CSV data before creating items or updating products
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST method allowed'})
    
    try:
        platform = request.POST.get('platform')
        csv_file = request.FILES.get('csv_file')
        operation_type = request.POST.get('operation_type', 'bulk_creation')  # Default to bulk creation
        
        if not platform or not csv_file:
            return JsonResponse({'success': False, 'message': 'Platform and CSV file are required'})
        
        # Process CSV file for preview
        import csv
        import io
        
        # Read CSV content with encoding fallback
        csv_content, encoding_used = decode_csv_upload(csv_file)
        csv_reader = csv.DictReader(io.StringIO(csv_content))

        preview_data = []
        errors = []
        warnings = []

        # Header validation based on operation type
        if operation_type == 'product_update':
            # Product update: only these 5 fields allowed
            required_headers = {'item_code', 'units'}
            optional_headers = {'mrp', 'cost', 'stock'}
            allowed_headers = required_headers | optional_headers
        else:
            # Bulk creation: requires all item creation fields
            required_headers = {'wrap', 'item_code', 'description', 'units', 'sku', 'pack_description'}
            optional_headers = {'barcode', 'mrp', 'selling_price', 'cost', 'stock', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty'}
            allowed_headers = required_headers | optional_headers
        
        # Filter out empty header fields (from trailing delimiters)
        from .utils import normalize_csv_header
        header_fields = [normalize_csv_header(h) for h in (csv_reader.fieldnames or []) if h and h.strip()]
        if not header_fields:
            return JsonResponse({'success': False, 'message': 'CSV is missing header row. Include headers exactly as specified.'})
        if 'is_active' in header_fields:
            return JsonResponse({'success': False, 'message': "Column 'is_active' is not allowed in the CSV. Remove it and try again."})
        missing_required = sorted(list(required_headers - set(header_fields)))
        unknown_headers = sorted([h for h in header_fields if h and h not in allowed_headers])
        if missing_required:
            return JsonResponse({'success': False, 'message': f"Missing required columns: {', '.join(missing_required)}. The file was rejected."})
        if unknown_headers:
            return JsonResponse({'success': False, 'message': f"Unknown columns present: {', '.join(unknown_headers)}. Only defined headers are allowed."})

        # Read CSV rows for preview
        csv_rows = list(csv_reader)
        
        # Detect operation type based on CSV headers if not specified
        if csv_rows and operation_type == 'bulk_creation':
            headers = list(csv_rows[0].keys())
            # Check if this looks like a product update CSV (has Item Code, Units, MRP, Stock)
            product_update_headers = ['Item Code', 'Units', 'MRP', 'Stock']
            if all(header in headers for header in product_update_headers):
                operation_type = 'product_update'
        
        # Validate entire file rows strictly; reject on any missing required or invalid numeric
        if operation_type == 'bulk_creation':
            fatal_row_errors = []
            for idx, r in enumerate(csv_rows, start=2):
                m = {
                    'wrap': r.get('wrap', '').strip(),
                    'item_code': r.get('item_code', '').strip(),
                    'description': r.get('description', '').strip(),
                    'units': r.get('units', '').strip(),
                    'sku': r.get('sku', '').strip(),
                    'pack_description': r.get('pack_description', '').strip(),
                }
                missing = [k for k, v in m.items() if not v]
                if missing:
                    fatal_row_errors.append(f"Row {idx}: Missing mandatory fields: {', '.join(missing)}")
                    continue
                # Validate wrap strictly
                if m['wrap'] not in ALLOWED_WRAP_VALUES:
                    fatal_row_errors.append(f"Row {idx}: Wrap must be 9900 or 10000 (got '{m['wrap']}')")
                    continue
                # Optional numeric validations
                try:
                    sp = r.get('selling_price', '').strip()
                    st = r.get('stock', '').strip()
                    c = r.get('cost', '').strip()
                    mval = r.get('mrp', '').strip()
                    wdf = r.get('weight_division_factor', '').strip()
                    ocq = r.get('outer_case_quantity', '').strip()
                    minq = r.get('minimum_qty', '').strip()
                    if sp:
                        Decimal(sp)
                    if st:
                        int(st)
                    if c:
                        Decimal(c)
                    if mval:
                        Decimal(mval)
                    if wdf:
                        Decimal(wdf)
                    if ocq:
                        int(ocq)
                    if minq:
                        int(minq)
                except (InvalidOperation, ValueError):
                    fatal_row_errors.append(f"Row {idx}: Invalid numeric values in one of [selling_price, stock, cost, mrp, weight_division_factor, outer_case_quantity, minimum_qty]")
                    continue
            if fatal_row_errors:
                # Return a concise message; front-end shows only message when success=false
                first = fatal_row_errors[:3]
                return JsonResponse({
                    'success': False,
                    'message': 'CSV validation failed; the file was rejected. First issues: ' + ' | '.join(first)
                })

        for row_num, row in enumerate(csv_rows[:20], start=2):  # Preview first 20 rows
            try:
                if operation_type == 'product_update':
                    # Product Update CSV validation
                    item_code = row.get('Item Code', row.get('item_code', '')).strip()
                    units = row.get('Units', row.get('units', '')).strip()
                    mrp_str = row.get('MRP', row.get('mrp', '')).strip()
                    cost_str = row.get('Cost', row.get('cost', '')).strip()
                    stock_str = row.get('Stock', row.get('stock', '')).strip()
                    
                    # Validate mandatory fields for product update (only item_code and units required)
                    row_status = 'valid'
                    row_errors = []
                    
                    if not item_code:
                        row_errors.append("Missing: item_code")
                        row_status = 'error'
                    if not units:
                        row_errors.append("Missing: units")
                        row_status = 'error'
                    
                    # Validate numeric fields
                    mrp = 0.0
                    cost = 0.0
                    stock = 0
                    try:
                        if mrp_str:
                            mrp = float(mrp_str)
                        if cost_str:
                            cost = float(cost_str)
                        if stock_str:
                            stock = int(float(stock_str))  # Handle decimal stock values
                    except ValueError:
                        row_errors.append("Invalid numeric values")
                        row_status = 'error'
                    
                    # Check if item exists for product update - PLATFORM ISOLATED
                    from .models import Item
                    if item_code and units:
                        if not Item.objects.filter(item_code=item_code, units=units, platform=platform).exists():
                            row_errors.append(f"Item '{item_code}' ({units}) not found in {platform.title()} platform")
                            row_status = 'error'
                    
                    preview_data.append({
                        'row_number': row_num,
                        'status': row_status,
                        'errors': row_errors,
                        'data': {
                            'item_code': item_code,
                            'units': units,
                            'mrp': mrp,
                            'cost': cost,
                            'stock': stock
                        }
                    })
                    
                else:
                    # Bulk Item Creation CSV validation (original logic)
                    base_item_code = row.get('item_code', '').strip()
                    base_sku = row.get('sku', '').strip()
                    
                    # Validate mandatory fields
                    mandatory_fields = {
                        'wrap': row.get('wrap', '').strip(),
                        'item_code': base_item_code,
                        'description': row.get('description', '').strip(),
                        'units': row.get('units', '').strip(),
                        'sku': base_sku,
                        'pack_description': row.get('pack_description', '').strip()
                    }
                    
                    # Handle optional fields
                    selling_price_str = row.get('selling_price', '').strip()
                    stock_str = row.get('stock', '').strip()
                    cost_str = row.get('cost', '').strip()
                    mrp_str = row.get('mrp', '').strip()
                    wdf_str = row.get('weight_division_factor', '').strip()
                    ocq_str = row.get('outer_case_quantity', '').strip()
                    minq_str = row.get('minimum_qty', '').strip()
                    
                    # Check for missing mandatory fields
                    missing_fields = [field for field, value in mandatory_fields.items() if not value]
                    row_status = 'valid'
                    row_errors = []
                    
                    if missing_fields:
                        row_errors.append(f"Missing: {', '.join(missing_fields)}")
                        row_status = 'error'
                    
                    # Validate numeric fields (optional fields are now weight_division_factor, outer_case_quantity, minimum_qty)
                    try:
                        selling_price = float(selling_price_str) if selling_price_str else 0.0
                        stock = int(stock_str) if stock_str else 0
                        cost = float(cost_str) if cost_str else 0.0
                        mrp = float(mrp_str) if mrp_str else 0.0
                        wdf = float(wdf_str) if wdf_str else None
                        ocq = int(ocq_str) if ocq_str else None
                        minq = int(minq_str) if minq_str else None
                    except ValueError:
                        row_errors.append("Invalid numeric values in one of [selling_price, stock, cost, mrp, weight_division_factor, outer_case_quantity, minimum_qty]")
                        row_status = 'error'
                        selling_price = 0.0
                        stock = 0
                        cost = 0.0
                        mrp = 0.0
                        wdf = None
                        ocq = None
                        minq = None
                    
                    # Handle optional fields
                    barcode = row.get('barcode', '').strip()
                    
                    # Check if item already exists and apply platform-specific duplicate handling
                    from .models import Item, ItemOutlet
                    existing_item = Item.objects.filter(sku=base_sku).first()
                    if existing_item:
                        if platform in ('pasons', 'talabat'):
                            linked = ItemOutlet.objects.filter(
                                item=existing_item,
                                outlet__platforms=platform  # STRICT isolation
                            ).exists()
                            if linked:
                                # Duplicate within selected platform: mark as row error
                                row_errors.append(
                                    f"SKU '{base_sku}' already exists for {platform}; duplicate creation is not allowed."
                                )
                                row_status = 'error'
                            else:
                                # Cross-platform reuse: no warning; item will be linked on upload
                                pass
                        else:
                            # Unknown platform context: keep as valid without warnings
                            pass
                    
                    preview_data.append({
                        'row_number': row_num,
                        'status': row_status,
                        'errors': row_errors,
                        'data': {
                            'wrap': mandatory_fields['wrap'],
                            'item_code': base_item_code,
                            'description': mandatory_fields['description'],
                            'pack_description': mandatory_fields['pack_description'],
                            'units': mandatory_fields['units'],
                            'barcode': barcode,
                            'sku': base_sku,
                            'selling_price': selling_price,
                            'stock': stock,
                            'cost': cost,
                            'mrp': mrp,
                            'weight_division_factor': wdf,
                            'outer_case_quantity': ocq,
                            'minimum_qty': minq
                        }
                    })
                
            except Exception as e:
                errors.append(f"Row {row_num}: Error processing - {str(e)}")
        
        # Get platform outlets info
        from .models import Outlet
        from django.db import models
        outlets = Outlet.objects.filter(
            is_active=True,
            platforms=platform  # STRICT isolation
        )
        
        return JsonResponse({
            'success': True,
            'preview_data': preview_data,
            'total_rows': len(csv_rows),
            'preview_rows': len(preview_data),
            'errors': errors,
            'warnings': warnings,
            'platform': platform,
            'outlets': [{'name': outlet.name, 'id': outlet.id} for outlet in outlets],
            'encoding_used': encoding_used
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing CSV: {str(e)}'})


@login_required
@rate_limit(max_requests=10, time_window_seconds=60)  # 10 requests per minute
def delete_items_api(request):
    """
    API endpoint for deleting items (both bulk and single deletion)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST method allowed'})
    
    try:
        import json
        from html import unescape as html_unescape
        from .models import Item
        from django.db import models
        
        data = json.loads(request.body)
        combination_keys = data.get('combination_keys', [])  # item_code|description|sku
        deletion_type = data.get('deletion_type', 'single')
        delete_scope = data.get('delete_scope', 'selected')  # selected | current_page | filtered | all
        platform = (data.get('platform') or '').strip()

        from django.db.models import Q, Count

        # Build queryset to delete based on scope
        items_to_delete = Item.objects.none()
        scope_description = ''

        if delete_scope == 'selected':
            if not combination_keys:
                return JsonResponse({'success': False, 'message': 'No items selected for deletion'})

            query = Q()
            for combination_key in combination_keys:
                parts = combination_key.split('|')
                if len(parts) != 3:
                    continue
                item_code, description, sku = parts
                # Unescape HTML entities in description (e.g., &#x27; -> ')
                description = html_unescape(description)
                query |= Q(
                    item_code__iexact=item_code,
                    description__iexact=description,
                    sku__iexact=sku
                )
            items_to_delete = Item.objects.filter(query)
            if platform in ('pasons', 'talabat'):
                items_to_delete = items_to_delete.filter(platform=platform)
            scope_description = f"selected ({len(combination_keys)})"

        elif delete_scope == 'current_page':
            # Expect combination_keys of the currently displayed items (100/200)
            if not combination_keys:
                return JsonResponse({'success': False, 'message': 'No page items provided for deletion'})
            query = Q()
            for combination_key in combination_keys:
                parts = combination_key.split('|')
                if len(parts) != 3:
                    continue
                item_code, description, sku = parts
                # Unescape HTML entities in description (e.g., &#x27; -> ')
                description = html_unescape(description)
                query |= Q(
                    item_code__iexact=item_code,
                    description__iexact=description,
                    sku__iexact=sku
                )
            items_to_delete = Item.objects.filter(query)
            if platform in ('pasons', 'talabat'):
                items_to_delete = items_to_delete.filter(platform=platform)
            scope_description = f"current_page ({len(combination_keys)})"

        elif delete_scope == 'filtered':
            # Filters payload matching search_product_api params
            filters = data.get('filters', {})
            platform = (data.get('platform') or '').strip()
            include_inactive = (data.get('include_inactive') or False) in (True, '1', 'true', 'True')
            item_code = (filters.get('item_code') or '').strip()
            description = (filters.get('description') or '').strip()
            barcode = (filters.get('barcode') or '').strip()
            sku = (filters.get('sku') or '').strip()
            price_min = (filters.get('price_min') or '').strip()
            price_max = (filters.get('price_max') or '').strip()
            stock_min = (filters.get('stock_min') or '').strip()
            stock_max = (filters.get('stock_max') or '').strip()

            qs = Item.objects.all() if include_inactive else Item.objects.filter(is_active=True)
            if platform in ('pasons', 'talabat'):
                qs = qs.filter(platform=platform)
            if item_code:
                qs = qs.filter(item_code__icontains=item_code)
            if description:
                qs = qs.filter(description__icontains=description)
            if barcode:
                qs = qs.filter(barcode__icontains=barcode)
            if sku:
                qs = qs.filter(sku__icontains=sku)
            if price_min:
                try:
                    qs = qs.filter(selling_price__gte=float(price_min))
                except ValueError:
                    pass
            if price_max:
                try:
                    qs = qs.filter(selling_price__lte=float(price_max))
                except ValueError:
                    pass
            if stock_min:
                try:
                    qs = qs.filter(stock__gte=int(stock_min))
                except ValueError:
                    pass
            if stock_max:
                try:
                    qs = qs.filter(stock__lte=int(stock_max))
                except ValueError:
                    pass

            items_to_delete = qs
            scope_description = "filtered"

        elif delete_scope == 'all':
            # Require strong confirmation to prevent accidental wipe
            confirm_all = data.get('confirm_all', False)
            confirm_text = (data.get('confirm_text') or '').strip()
            if not (confirm_all and confirm_text.upper() == 'DELETE ALL'):
                return JsonResponse({'success': False, 'message': 'Full delete requires confirmation: type "DELETE ALL"'})
            # If a platform is provided, limit to items linked to that platform; otherwise entire database
            if (data.get('platform') or '').strip() in ('pasons', 'talabat'):
                platform = (data.get('platform') or '').strip()
                items_to_delete = Item.objects.filter(platform=platform)
                scope_description = f"entire_database_for_platform_{platform}"
            else:
                items_to_delete = Item.objects.all()
                scope_description = "entire_database"

        else:
            return JsonResponse({'success': False, 'message': 'Invalid delete scope'})

        if not items_to_delete.exists():
            return JsonResponse({'success': False, 'message': 'No items found for deletion'})

        # For platform-safe deletion: first remove ItemOutlet associations for selected platform,
        # then delete Item objects that become orphaned (no outlets remain)
        # Collect brief audit info (limit to first 50 to avoid log blow-up)
        deleted_items_info = []
        for item in items_to_delete[:50]:
            deleted_items_info.append({
                'id': item.id,
                'item_code': item.item_code,
                'description': item.description,
                'sku': item.sku
            })

        # Delete associations specific to selected platform if provided; otherwise, delete all associations
        if platform in ('pasons', 'talabat'):
            associations_qs = ItemOutlet.objects.filter(
                item__in=items_to_delete,
                outlet__platforms=platform
            )
        else:
            associations_qs = ItemOutlet.objects.filter(item__in=items_to_delete)

        associations_count = associations_qs.count()
        associations_qs.delete()

        # Now delete items that are no longer associated with any outlet
        orphan_items_qs = Item.objects.filter(id__in=items_to_delete.values_list('id', flat=True))\
            .annotate(outlet_count=Count('item_outlets')).filter(outlet_count=0)
        items_deleted_count = orphan_items_qs.count()
        orphan_items_qs.delete()

        logger.info(
            f"User {request.user.username} deletion via {deletion_type} ({scope_description}) | Associations removed: {associations_count}, Items deleted: {items_deleted_count}"
        )

        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {items_deleted_count} item(s) and removed {associations_count} association(s)',
            'deleted_count': items_deleted_count,
            'associations_deleted': associations_count,
            'deletion_type': deletion_type,
            'delete_scope': delete_scope
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Error in delete_items_api: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error deleting items: {str(e)}'})


@login_required
def export_feed_api(request):
    """
    API endpoint to export production data (SKU, selling_price, stock_status) for e-commerce sync.
    
    FEATURES:
    ✓ Full export: All items at outlet
    ✓ Partial export: Only items changed since last successful export
    ✓ Data validation: All items checked before export
    ✓ Integrity tracking: ExportHistory records every export
    ✓ Transaction safety: Atomic operations
    
    Query parameters:
    - platform: 'pasons' or 'talabat' (required)
    - outlet_id: ID of outlet to export (required)
    - export_type: 'full' or 'partial' (optional, default='partial' for auto-detect)
    
    Returns:
    - Success: CSV file download with headers [sku, selling_price, stock_status]
    - Error: JSON response with error details
    
    CSV FILENAME FORMAT:
        OutletName_YYYY-MM-DD_HHMMSS.csv
        Example: PASONS_14_2025-12-15_143530.csv
    
    EXPORT LOGIC:
    - Full: Export all active items at outlet
    - Partial: Export only items changed since last successful full/partial export
    - First export: Always treated as full export
    """
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from .export_service import ExportService
    
    # Parse query parameters
    platform = request.GET.get('platform', '').strip()
    outlet_id = request.GET.get('outlet_id', '').strip()
    export_type_param = request.GET.get('export_type', '').strip()  # Can be '', 'full', or 'partial'
    exclude_promotions = request.GET.get('exclude_promotions', '').strip().lower() in ('true', '1', 'yes')
    
    # VALIDATION 1: Platform
    if platform not in ('pasons', 'talabat'):
        logger.warning(f"Invalid platform: {platform}")
        return JsonResponse({
            'success': False,
            'message': 'Invalid platform. Must be "pasons" or "talabat".'
        })
    
    # VALIDATION 2: Outlet ID
    if not outlet_id or not outlet_id.isdigit():
        logger.warning(f"Invalid outlet_id: {outlet_id}")
        return JsonResponse({
            'success': False,
            'message': 'Outlet ID is required and must be numeric.'
        })
    
    try:
        outlet = Outlet.objects.get(id=int(outlet_id), platforms=platform)
    except Outlet.DoesNotExist:
        logger.warning(f"Outlet not found: id={outlet_id}, platform={platform}")
        return JsonResponse({
            'success': False,
            'message': f'Outlet with ID {outlet_id} not found on {platform} platform.'
        })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Outlet ID must be numeric.'
        })
    
    try:
        # Create export service
        export_service = ExportService(outlet, platform)
        
        # Determine export type
        # If user explicitly requested full/partial, use that; otherwise auto-detect
        manual_export_type = export_type_param if export_type_param in ('full', 'partial') else None
        
        # Execute export
        export_data, export_history = export_service.export(
            user=request.user if request.user.is_authenticated else None,
            manual_export_type=manual_export_type,
            exclude_promotions=exclude_promotions
        )
        
        # Check if export succeeded
        if export_data is None:
            logger.error(
                f"Export failed for {outlet.name}: {export_history.status}. "
                f"Errors: {export_history.validation_errors}"
            )
            # Parse validation errors from JSON string
            import json
            validation_errors_list = []
            if export_history.validation_errors:
                try:
                    validation_errors_list = json.loads(export_history.validation_errors)
                except (json.JSONDecodeError, TypeError):
                    validation_errors_list = [export_history.validation_errors]
            
            return JsonResponse({
                'success': False,
                'message': f'Export validation failed: {export_history.get_status_display()}',
                'validation_errors': validation_errors_list,
                'export_history_id': export_history.id
            })
        
        # EXPORT SUCCEEDED - Generate CSV
        now = timezone.localtime(timezone.now())
        outlet_name_clean = outlet.name.replace(' ', '-').replace('/', '-')
        filename = f"{outlet_name_clean}-{now.strftime('%Y-%m-%d-%H%M%S')}.csv"
        
        # Save CSV file to disk for re-download
        import os
        from django.conf import settings
        
        # Create exports directory if it doesn't exist
        exports_dir = getattr(settings, 'EXPORT_FILES_DIR', settings.MEDIA_ROOT / 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        file_path = os.path.join(exports_dir, filename)
        
        # Write CSV to file - different format for Pasons vs Talabat
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            file_writer = csv.writer(f)  # Use comma delimiter for both platforms
            
            if platform == 'talabat':
                # Talabat format: barcode, sku, reason, start_date, end_date, campaign_status, discounted_price, max_no_of_orders, price, active
                file_writer.writerow(['barcode', 'sku', 'reason', 'start_date', 'end_date', 'campaign_status', 'discounted_price', 'max_no_of_orders', 'price', 'active'])
                for row in export_data:
                    file_writer.writerow([
                        row['barcode'],           # barcode
                        row['sku'],               # sku
                        '',                       # reason (placeholder)
                        '',                       # start_date (placeholder)
                        '',                       # end_date (placeholder)
                        '',                       # campaign_status (placeholder)
                        '',                       # discounted_price (placeholder)
                        '',                       # max_no_of_orders (placeholder)
                        row['selling_price'],     # price
                        row['stock_status']       # active (stock_status)
                    ])
            else:
                # Pasons format: sku, selling_price, stock_status, availability_status
                file_writer.writerow(['sku', 'selling_price', 'stock_status', 'availability_status'])
                for row in export_data:
                    file_writer.writerow([
                        row['sku'],
                        row['selling_price'],
                        row['stock_status'],
                        row['stock_status']
                    ])
        
        # Update ExportHistory with filename
        export_history.file_name = filename
        export_history.save(update_fields=['file_name'])
        
        # Create CSV response for immediate download
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        if platform == 'talabat':
            writer = csv.writer(response)  # Use comma delimiter for Excel compatibility
            # Talabat format
            writer.writerow(['barcode', 'sku', 'reason', 'start_date', 'end_date', 'campaign_status', 'discounted_price', 'max_no_of_orders', 'price', 'active'])
            for row in export_data:
                writer.writerow([
                    row['barcode'],
                    row['sku'],
                    '',  # reason
                    '',  # start_date
                    '',  # end_date
                    '',  # campaign_status
                    '',  # discounted_price
                    '',  # max_no_of_orders
                    row['selling_price'],
                    row['stock_status']
                ])
        else:
            writer = csv.writer(response)
            # Pasons format
            writer.writerow(['sku', 'selling_price', 'stock_status', 'availability_status'])
            for row in export_data:
                writer.writerow([
                    row['sku'],
                    row['selling_price'],
                    row['stock_status'],
                    row['stock_status']
                ])
        
        logger.info(
            f"Export successful: {outlet.name} ({platform}) - "
            f"{export_history.get_export_type_display()} - "
            f"{len(export_data)} items - File: {filename} (saved to disk)"
        )
        
        return response
    
    except Exception as e:
        logger.exception(f"Unexpected error in export_feed_api: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An unexpected error occurred during export.',
            'error_details': str(e)
        })


@login_required
def download_export_file(request):
    """
    Download a previously exported CSV file by filename.
    This serves the stored file instead of re-generating the export.
    """
    import os
    from django.conf import settings
    from django.http import FileResponse, Http404
    
    filename = request.GET.get('filename', '').strip()
    
    if not filename:
        return JsonResponse({'success': False, 'message': 'Filename is required'})
    
    # Security: Only allow .csv files and prevent directory traversal
    if not filename.endswith('.csv') or '/' in filename or '\\' in filename or '..' in filename:
        return JsonResponse({'success': False, 'message': 'Invalid filename'})
    
    # Get exports directory
    exports_dir = getattr(settings, 'EXPORT_FILES_DIR', settings.MEDIA_ROOT / 'exports')
    file_path = os.path.join(exports_dir, filename)
    
    if not os.path.exists(file_path):
        return JsonResponse({'success': False, 'message': 'File not found. It may have been deleted.'})
    
    # Serve the file
    response = FileResponse(open(file_path, 'rb'), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def erp_export_api(request):
    """
    ERP Export API for Talabat platform.
    
    CSV FORMAT: Party, Item Code, Location, Unit, Price
    - Party: Fixed value "DT0072"
    - Item Code: item.item_code
    - Location: Placeholder (empty)
    - Unit: item.units
    - Price: Converted selling price
        - wrap=9900: selling_price * WDF (weight_division_factor)
        - wrap=10000: selling_price (no conversion)
    
    Supports full and partial export with delta tracking.
    """
    import csv
    import os
    from django.conf import settings
    from django.http import HttpResponse
    from django.utils import timezone
    from decimal import Decimal
    from .models import ERPExportHistory
    
    try:
        # Get parameters
        outlet_id = request.GET.get('outlet_id', '').strip()
        export_type = request.GET.get('export_type', 'full').strip()
        exclude_promotions = request.GET.get('exclude_promotions', '').strip().lower() in ('true', '1', 'yes')
        
        if not outlet_id:
            return JsonResponse({'success': False, 'message': 'outlet_id is required'})
        
        if export_type not in ('full', 'partial'):
            return JsonResponse({'success': False, 'message': 'export_type must be "full" or "partial"'})
        
        # Get outlet (must be Talabat)
        try:
            outlet = Outlet.objects.get(id=outlet_id, platforms='talabat', is_active=True)
        except Outlet.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Outlet not found or not a Talabat outlet'})
        
        # Create export history record
        erp_export = ERPExportHistory.objects.create(
            outlet=outlet,
            export_type=export_type,
            status='success',
            created_by=request.user if request.user.is_authenticated else None
        )
        
        # Get items for this outlet
        item_outlets = ItemOutlet.objects.filter(
            outlet=outlet,
            item__platform='talabat',
            is_active_in_outlet=True
        ).select_related('item')
        
        # Filter out promotion items if requested
        if exclude_promotions:
            item_outlets = item_outlets.filter(is_on_promotion=False)
            logger.info(f"ERP Export: Excluding promotion items")
        
        # For partial export, compare with last export
        if export_type == 'partial':
            last_export = ERPExportHistory.get_latest_successful_export(outlet)
            if last_export:
                # FIXED: Use value-based comparison like shop integration
                # Compare current values vs last exported values for ALL items (not just recently updated)
                logger.info(f"Partial export: comparing current values vs last exported values (last export: {last_export.export_timestamp})")
                
                # Get ALL active items and compare current vs exported values
                all_items = item_outlets.select_related('item')
                
                changed_items = []
                for io in all_items:
                    # Calculate current ERP price
                    current_erp_price = Decimal(str(calculate_erp_price(io, io.item)))
                    
                    # Compare with last exported price
                    exported_price = io.erp_export_price or Decimal('0')
                    
                    # Include if price changed OR never exported before
                    if current_erp_price != exported_price:
                        logger.debug(f"Item {io.item.item_code} marked for export: price: {exported_price} → {current_erp_price}")
                        changed_items.append(io.id)
                
                # Filter to only changed items
                if changed_items:
                    item_outlets = item_outlets.filter(id__in=changed_items)
                    logger.info(f"Partial export: {len(changed_items)} items have price changes")
                else:
                    # No changes found, return empty
                    logger.info("Partial export: No items with price changes found")
                    item_outlets = item_outlets.none()
                
                # If no changes found, return empty queryset
                if not item_outlets.exists():
                    logger.info("Partial export: No changes detected since last export")
            else:
                # No previous export found - treat as full export
                logger.info("Partial export: No previous export found, performing full export")
        
        # Build export data
        raw_export_data = []
        for io in item_outlets:
            item = io.item
            # Calculate converted price based on wrap type
            converted_price = calculate_erp_price(io, item)
            
            raw_export_data.append({
                'party': 'DT0072',
                'item_code': item.item_code,
                'location': '',  # Placeholder
                'unit': item.units,
                'price': converted_price,
                'wrap': str(item.wrap)
            })
        
        # Remove duplicates for wrap=9900: same item_code + units → keep LOWEST price
        # wrap=10000 items: include all (no duplicates expected)
        export_data = []
        seen_keys = {}  # key = (item_code, unit) → lowest price row
        
        for row in raw_export_data:
            if row['wrap'] == '9900':
                # For wrap=9900, deduplicate by (item_code, unit), keep lowest price
                key = (row['item_code'], row['unit'])
                if key not in seen_keys or row['price'] < seen_keys[key]['price']:
                    seen_keys[key] = row
            else:
                # wrap=10000 or other: include all
                export_data.append(row)
        
        # Add deduplicated wrap=9900 items
        for row in seen_keys.values():
            export_data.append(row)
        
        # Remove 'wrap' key from final export (not needed in CSV)
        for row in export_data:
            row.pop('wrap', None)
        
        # Update export history
        erp_export.item_count = len(export_data)
        
        if len(export_data) == 0:
            erp_export.status = 'success'
            erp_export.file_name = ''
            erp_export.save()
            
            # Log why no data was exported
            if export_type == 'partial':
                logger.info(f"Partial export: No changes detected for outlet {outlet.name}")
            else:
                logger.warning(f"Full export: No active items found for outlet {outlet.name}")
            
            return JsonResponse({
                'success': True,
                'message': f'No items to export ({export_type} export - no changes detected)' if export_type == 'partial' else 'No active items found for export',
                'item_count': 0
            })
        
        # Generate filename
        now = timezone.localtime(timezone.now())
        outlet_name_clean = outlet.name.replace(' ', '-').replace('/', '-')
        filename = f"ERP-{outlet_name_clean}-{now.strftime('%Y-%m-%d-%H%M%S')}.csv"
        
        # Save CSV file to disk
        exports_dir = getattr(settings, 'EXPORT_FILES_DIR', settings.MEDIA_ROOT / 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        file_path = os.path.join(exports_dir, filename)
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            file_writer = csv.writer(f)
            file_writer.writerow(['Party', 'Item Code', 'Location', 'Unit', 'Price'])
            for row in export_data:
                file_writer.writerow([
                    row['party'],
                    row['item_code'],
                    row['location'],
                    row['unit'],
                    row['price']
                ])
        
        # Update export history
        erp_export.file_name = filename
        erp_export.save()
        
        # Update tracking fields for partial export (OPTIMIZED)
        # Only update items that were actually exported, not all items
        if export_type == 'partial' and len(export_data) > 0:
            # Get the item_outlets that were actually exported
            exported_item_codes = [row['item_code'] for row in export_data]
            item_outlets_to_update = ItemOutlet.objects.filter(
                outlet=outlet,
                item__platform='talabat',
                item__item_code__in=exported_item_codes,
                is_active_in_outlet=True
            ).select_related('item')
            
            # Batch update with calculated prices
            updates = []
            for io in item_outlets_to_update:
                io.erp_export_price = Decimal(str(calculate_erp_price(io, io.item)))
                updates.append(io)
            
            if updates:
                ItemOutlet.objects.bulk_update(
                    updates, 
                    ['erp_export_price'],
                    batch_size=500
                )
                logger.info(f"Updated tracking for {len(updates)} exported items")
        elif export_type == 'full':
            # For full export, update all active items
            item_outlets_to_update = ItemOutlet.objects.filter(
                outlet=outlet,
                item__platform='talabat',
                is_active_in_outlet=True
            ).select_related('item')
            
            updates = []
            for io in item_outlets_to_update:
                io.erp_export_price = Decimal(str(calculate_erp_price(io, io.item)))
                updates.append(io)
            
            if updates:
                ItemOutlet.objects.bulk_update(
                    updates, 
                    ['erp_export_price'],
                    batch_size=500
                )
                logger.info(f"Updated tracking for {len(updates)} items (full export)")
        
        logger.info(f"ERP Export successful: {outlet.name} - {len(export_data)} items - File: {filename}")
        
        # Return JSON success response (no download - user downloads from history)
        return JsonResponse({
            'success': True,
            'message': f'Export generated successfully. {len(export_data)} items exported.',
            'filename': filename,
            'item_count': len(export_data)
        })
    
    except Exception as e:
        logger.exception(f"ERP Export error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred during ERP export.',
            'error_details': str(e)
        })


def calculate_erp_price(item_outlet, item):
    """
    Calculate ERP price based on wrap type.
    
    - wrap=9900: selling_price * WDF (weight_division_factor)
    - wrap=10000: selling_price (no conversion)
    """
    selling_price = item_outlet.outlet_selling_price or item.selling_price or Decimal('0')
    
    if str(item.wrap) == '9900':
        # Multiply by WDF for wrap=9900
        wdf = item.weight_division_factor or 1
        converted_price = float(selling_price) * float(wdf)
    else:
        # wrap=10000 or other - no conversion
        converted_price = float(selling_price)
    
    return round(converted_price, 2)


@login_required
def download_erp_export_file(request):
    """
    Download a previously exported ERP CSV file by filename.
    """
    import os
    from django.conf import settings
    from django.http import FileResponse
    
    filename = request.GET.get('filename', '').strip()
    
    if not filename:
        return JsonResponse({'success': False, 'message': 'Filename is required'})
    
    # Security: Only allow .csv files and prevent directory traversal
    if not filename.endswith('.csv') or '/' in filename or '\\' in filename or '..' in filename:
        return JsonResponse({'success': False, 'message': 'Invalid filename'})
    
    # Get exports directory
    exports_dir = getattr(settings, 'EXPORT_FILES_DIR', settings.MEDIA_ROOT / 'exports')
    file_path = os.path.join(exports_dir, filename)
    
    if not os.path.exists(file_path):
        return JsonResponse({'success': False, 'message': 'File not found. It may have been deleted.'})
    
    # Serve the file
    response = FileResponse(open(file_path, 'rb'), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ============================================================================
# REPORTS & DATA EXPORT VIEWS
# ============================================================================

@login_required
def reports_page(request):
    """
    Reports page - displays export options for items data.
    """
    context = {
        'page_title': 'Reports & Data Export',
        'active_nav': 'reports',
    }
    return render(request, 'reports.html', context)


@login_required
def export_all_items(request):
    """
    Export ALL items from database (all platforms, linked or not linked to outlets).
    
    CSV Fields: item_code, units, description, pack_description, sku, barcode, wrap,
                weight_division_factor, outer_case_quantity, minimum_qty, platform,
                talabat_margin, is_active
    """
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from .models import Item
    
    if request.method != 'POST':
        return redirect('integration:reports')
    
    # Generate filename with timestamp
    timestamp = timezone.localtime().strftime('%Y-%m-%d-%H%M%S')
    filename = f'all-items-export-{timestamp}.csv'
    
    # Create HTTP response with CSV content type
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header row
    writer.writerow([
        'item_code', 'units', 'description', 'pack_description', 'sku', 'barcode',
        'wrap', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty',
        'platform', 'talabat_margin', 'is_active'
    ])
    
    # Fetch all items
    items = Item.objects.all().order_by('platform', 'item_code')
    
    for item in items:
        writer.writerow([
            item.item_code,
            item.units,
            item.description,
            item.pack_description or '',
            item.sku,
            item.barcode or '',
            item.wrap or '',
            float(item.weight_division_factor) if item.weight_division_factor else '',
            item.outer_case_quantity or '',
            item.minimum_qty or '',
            item.platform,
            float(item.effective_talabat_margin) if item.platform == 'talabat' else '',
            'Yes' if item.is_active else 'No'
        ])
    
    return response


@login_required
def export_platform_items(request):
    """
    Export items for a specific platform (all items, linked or not linked to outlets).
    
    CSV Fields: item_code, units, description, pack_description, sku, barcode, wrap,
                weight_division_factor, outer_case_quantity, minimum_qty, talabat_margin, is_active
    """
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from django.contrib import messages
    from .models import Item
    
    if request.method != 'POST':
        return redirect('integration:reports')
    
    platform = request.POST.get('platform', '').strip()
    
    if not platform or platform not in ('pasons', 'talabat'):
        messages.error(request, 'Please select a valid platform.')
        return redirect('integration:reports')
    
    # Generate filename with timestamp
    timestamp = timezone.localtime().strftime('%Y-%m-%d-%H%M%S')
    platform_name = 'Pasons' if platform == 'pasons' else 'Talabat'
    filename = f'{platform_name}-items-export-{timestamp}.csv'
    
    # Create HTTP response with CSV content type
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header row
    headers = [
        'item_code', 'units', 'description', 'pack_description', 'sku', 'barcode',
        'wrap', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty', 'is_active'
    ]
    
    # Add talabat_margin for Talabat platform
    if platform == 'talabat':
        headers.append('talabat_margin')
    
    writer.writerow(headers)
    
    # Fetch items for platform
    items = Item.objects.filter(platform=platform).order_by('item_code')
    
    for item in items:
        row = [
            item.item_code,
            item.units,
            item.description,
            item.pack_description or '',
            item.sku,
            item.barcode or '',
            item.wrap or '',
            float(item.weight_division_factor) if item.weight_division_factor else '',
            item.outer_case_quantity or '',
            item.minimum_qty or '',
            'Yes' if item.is_active else 'No'
        ]
        
        # Add talabat_margin for Talabat platform
        if platform == 'talabat':
            row.append(float(item.effective_talabat_margin) if item.effective_talabat_margin else '')
        
        writer.writerow(row)
    
    return response


@login_required
def export_outlet_items(request):
    """
    Export items linked to a specific outlet with outlet-specific data.
    
    CSV Fields: item_code, units, description, pack_description, sku, barcode, wrap,
                weight_division_factor, outer_case_quantity, minimum_qty, platform,
                talabat_margin, outlet_mrp, outlet_selling_price, outlet_stock, outlet_cost,
                price_locked, status_locked, is_active_in_outlet, stock_status
    """
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from django.contrib import messages
    from .models import Outlet, ItemOutlet
    
    if request.method != 'POST':
        return redirect('integration:reports')
    
    platform = request.POST.get('platform', '').strip()
    outlet_id = request.POST.get('outlet', '').strip()
    
    if not platform or platform not in ('pasons', 'talabat'):
        messages.error(request, 'Please select a valid platform.')
        return redirect('integration:reports')
    
    if not outlet_id:
        messages.error(request, 'Please select an outlet.')
        return redirect('integration:reports')
    
    try:
        outlet = Outlet.objects.get(id=outlet_id)
        
        # Validate outlet matches platform
        if outlet.platforms != platform:
            messages.error(request, f'Outlet "{outlet.name}" does not belong to {platform.title()} platform.')
            return redirect('integration:reports')
        
    except Outlet.DoesNotExist:
        messages.error(request, 'Selected outlet not found.')
        return redirect('integration:reports')
    
    # Generate filename with timestamp
    timestamp = timezone.localtime().strftime('%Y-%m-%d-%H%M%S')
    outlet_name_clean = outlet.name.replace(' ', '-').replace('/', '-')
    filename = f'{outlet_name_clean}-items-export-{timestamp}.csv'
    
    # Create HTTP response with CSV content type
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header row
    headers = [
        'item_code', 'units', 'description', 'pack_description', 'sku', 'barcode',
        'wrap', 'weight_division_factor', 'outer_case_quantity', 'minimum_qty',
        'platform', 'talabat_margin', 'outlet_mrp', 'outlet_selling_price',
        'outlet_stock', 'outlet_cost', 'converted_cost', 'is_active_in_outlet'
    ]
    writer.writerow(headers)
    
    # Fetch ItemOutlet records for this outlet with related Item data
    item_outlets = ItemOutlet.objects.filter(
        outlet=outlet
    ).select_related('item').order_by('item__item_code')
    
    for io in item_outlets:
        item = io.item
        
        writer.writerow([
            item.item_code,
            item.units,
            item.description,
            item.pack_description or '',
            item.sku,
            item.barcode or '',
            item.wrap or '',
            float(item.weight_division_factor) if item.weight_division_factor else '',
            item.outer_case_quantity or '',
            item.minimum_qty or '',
            item.platform,
            float(item.effective_talabat_margin) if item.platform == 'talabat' else '',
            float(io.outlet_mrp) if io.outlet_mrp else '',
            float(io.outlet_selling_price) if io.outlet_selling_price else '',
            io.outlet_stock or 0,
            float(io.outlet_cost) if io.outlet_cost else '',
            float(item.converted_cost) if item.converted_cost else '',
            'Active' if io.is_active_in_outlet else 'Inactive'
        ])
    
    return response



@login_required
def report_stats_api(request):
    """
    API endpoint to get report statistics for dashboard cards.
    """
    from .models import Item, Outlet, ItemOutlet
    
    try:
        # Platform-specific item counts
        pasons_items = Item.objects.filter(platform='pasons').count()
        talabat_items = Item.objects.filter(platform='talabat').count()
        
        # Platform-specific outlet counts
        pasons_outlets = Outlet.objects.filter(platforms='pasons', is_active=True).count()
        talabat_outlets = Outlet.objects.filter(platforms='talabat', is_active=True).count()
        
        return JsonResponse({
            'success': True,
            'stats': {
                'pasons_items': pasons_items,
                'talabat_items': talabat_items,
                'pasons_outlets': pasons_outlets,
                'talabat_outlets': talabat_outlets,
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e),
            'stats': {
                'pasons_items': 0,
                'talabat_items': 0,
                'pasons_outlets': 0,
                'talabat_outlets': 0,
            }
        })


@login_required
def report_data_api(request):
    """
    API endpoint to get report data for DataTables.
    
    Query params:
    - type: 'all', 'platform', 'outlet'
    - platform: 'pasons', 'talabat' (optional)
    - outlet: outlet_id (optional, for outlet type)
    - status: 'active', 'inactive' (optional)
    """
    from .models import Item, Outlet, ItemOutlet
    
    try:
        report_type = request.GET.get('type', 'all')
        platform = request.GET.get('platform', '').strip()
        outlet_id = request.GET.get('outlet', '').strip()
        status = request.GET.get('status', '').strip()
        
        headers = []
        rows = []
        
        if report_type == 'outlet' and outlet_id:
            # Outlet-specific report with ItemOutlet data - PLATFORM VALIDATION
            try:
                outlet = Outlet.objects.get(id=outlet_id)
            except Outlet.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Outlet not found'})
            
            # CRITICAL: Validate outlet platform matches requested platform
            if platform and outlet.platforms != platform:
                return JsonResponse({
                    'success': False, 
                    'message': f'Outlet "{outlet.name}" belongs to {outlet.platforms.title()} platform, not {platform.title()} platform.'
                })
            
            headers = [
                '#', 'Item Code', 'Description', 'Units', 'SKU', 'Barcode', 'Wrap',
                'WDF', 'OCQ', 'Min Qty', 'Platform', 'Outlet MRP', 'Outlet Selling Price',
                'Outlet Stock', 'Outlet Cost', 'Converted Cost', 'Active in Outlet'
            ]
            
            # Add Talabat Margin for Talabat platform
            if outlet.platforms == 'talabat':
                headers.insert(11, 'Talabat Margin')
            
            # PLATFORM ISOLATED: Only get ItemOutlets for items from the outlet's platform
            item_outlets = ItemOutlet.objects.filter(
                outlet=outlet,
                item__platform=outlet.platforms  # Ensure items match outlet platform
            ).select_related('item')
            
            # Apply status filter
            if status == 'active':
                item_outlets = item_outlets.filter(is_active_in_outlet=True)
            elif status == 'inactive':
                item_outlets = item_outlets.filter(is_active_in_outlet=False)
            
            item_outlets = item_outlets.order_by('item__item_code')
            
            for idx, io in enumerate(item_outlets, 1):
                item = io.item
                
                row = [
                    idx,
                    item.item_code,
                    item.description[:50] + '...' if len(item.description) > 50 else item.description,
                    item.units,
                    item.sku,
                    item.barcode or '-',
                    item.wrap or '-',
                    float(item.weight_division_factor) if item.weight_division_factor else '-',
                    item.outer_case_quantity or '-',
                    item.minimum_qty or '-',
                    item.platform.title(),
                ]
                
                # Add Talabat Margin for Talabat platform
                if outlet.platforms == 'talabat':
                    row.append(f"{float(item.effective_talabat_margin)}%" if item.effective_talabat_margin else '-')
                
                # Add outlet-specific fields
                row.extend([
                    float(io.outlet_mrp) if io.outlet_mrp else '-',
                    float(io.outlet_selling_price) if io.outlet_selling_price else '-',
                    io.outlet_stock or 0,
                    float(io.outlet_cost) if io.outlet_cost else '-',
                    float(item.converted_cost) if item.converted_cost else '-',
                    'Active' if io.is_active_in_outlet else 'Inactive'
                ])
                
                rows.append(row)
        
        else:
            # All items or platform-specific report - PLATFORM ISOLATION REQUIRED
            # CRITICAL: Always require platform filter to prevent data leakage
            if not platform or platform not in ('pasons', 'talabat'):
                return JsonResponse({
                    'success': False, 
                    'message': 'Platform filter is required. Please select Pasons or Talabat platform.'
                })
            
            headers = [
                '#', 'Item Code', 'Description', 'Units', 'SKU', 'Barcode', 'Wrap',
                'WDF', 'OCQ', 'Min Qty', 'Platform', 'Status'
            ]
            
            # PLATFORM ISOLATED: Only get items from selected platform
            items = Item.objects.filter(platform=platform)
            
            # Add Talabat Margin for Talabat platform
            if platform == 'talabat':
                headers.insert(11, 'Talabat Margin')
            
            # Apply status filter
            if status == 'active':
                items = items.filter(is_active=True)
            elif status == 'inactive':
                items = items.filter(is_active=False)
            
            items = items.order_by('platform', 'item_code')
            
            for idx, item in enumerate(items, 1):
                row = [
                    idx,
                    item.item_code,
                    item.description[:50] + '...' if len(item.description) > 50 else item.description,
                    item.units,
                    item.sku,
                    item.barcode or '-',
                    item.wrap or '-',
                    float(item.weight_division_factor) if item.weight_division_factor else '-',
                    item.outer_case_quantity or '-',
                    item.minimum_qty or '-',
                    item.platform.title(),
                ]
                
                # Add Talabat Margin for Talabat platform
                if platform == 'talabat':
                    row.append(f"{float(item.effective_talabat_margin)}%" if item.effective_talabat_margin else '-')
                
                row.append('Active' if item.is_active else 'Inactive')
                rows.append(row)
        
        # Get stats
        total_items = Item.objects.count()
        active_items = Item.objects.filter(is_active=True).count()
        outlets_count = Outlet.objects.filter(is_active=True).count()
        linked_items = ItemOutlet.objects.values('item_id').distinct().count()
        
        return JsonResponse({
            'success': True,
            'headers': headers,
            'rows': rows,
            'total_rows': len(rows),
            'stats': {
                'total_items': total_items,
                'active_items': active_items,
                'outlets_count': outlets_count,
                'linked_items': linked_items,
            }
        })
        
    except Exception as e:
        logger.error(f"Report data API error: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error loading report data: {str(e)}'
        })


@login_required
def locked_products_report(request):
    """
    Locked Products Report page - displays CLS and BLS locked items.
    """
    context = {
        'page_title': 'Locked Products Report',
        'active_nav': 'locked_products_report',
    }
    return render(request, 'locked_products_report.html', context)


@login_required
def locked_products_data_api(request):
    """
    API endpoint to get locked products data for DataTables.
    
    Query params:
    - platform: 'pasons', 'talabat' (required)
    - lock_type: 'cls_price', 'cls_status', 'bls_price', 'bls_status'
    - outlet: outlet_id (optional, for BLS reports)
    """
    from .models import Item, Outlet, ItemOutlet
    
    try:
        platform = request.GET.get('platform', '').strip()
        lock_type = request.GET.get('lock_type', '').strip()
        outlet_id = request.GET.get('outlet', '').strip()
        
        # Validate platform
        if not platform or platform not in ('pasons', 'talabat'):
            return JsonResponse({
                'success': False, 
                'message': 'Platform filter is required. Please select Pasons or Talabat platform.'
            })
        
        # Validate lock type
        if not lock_type or lock_type not in ('cls_price', 'cls_status', 'bls_price', 'bls_status'):
            return JsonResponse({
                'success': False, 
                'message': 'Lock type is required. Please select a valid lock type.'
            })
        
        headers = ['#', 'Item Code', 'Units', 'Description', 'Selling Price', 'Pack Description', 'Lock Type', 'Lock Level']
        rows = []
        
        if lock_type.startswith('cls'):
            # Central Locking System (CLS) - affects all outlets for the platform
            if lock_type == 'cls_price':
                items = Item.objects.filter(
                    platform=platform,
                    price_locked=True,
                    is_active=True
                ).order_by('item_code')
                lock_display = 'CLS Price Lock'
            else:  # cls_status
                items = Item.objects.filter(
                    platform=platform,
                    status_locked=True,
                    is_active=True
                ).order_by('item_code')
                lock_display = 'CLS Status Lock'
            
            for idx, item in enumerate(items, 1):
                rows.append([
                    idx,
                    item.item_code,
                    item.units,
                    item.description[:50] + '...' if len(item.description) > 50 else item.description,
                    float(item.selling_price) if item.selling_price else '-',
                    item.pack_description[:30] + '...' if item.pack_description and len(item.pack_description) > 30 else (item.pack_description or '-'),
                    lock_display,
                    'Central (All Outlets)'
                ])
        
        else:
            # Branch Locking System (BLS) - outlet-specific locks
            if not outlet_id:
                return JsonResponse({
                    'success': False, 
                    'message': 'Outlet selection is required for BLS reports.'
                })
            
            try:
                outlet = Outlet.objects.get(id=outlet_id, platforms=platform, is_active=True)
            except Outlet.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'Selected outlet not found or does not belong to the selected platform.'
                })
            
            # Add outlet name to headers
            headers.append('Outlet')
            
            if lock_type == 'bls_price':
                item_outlets = ItemOutlet.objects.filter(
                    outlet=outlet,
                    item__platform=platform,
                    price_locked=True,
                    is_active_in_outlet=True
                ).select_related('item').order_by('item__item_code')
                lock_display = 'BLS Price Lock'
            else:  # bls_status
                item_outlets = ItemOutlet.objects.filter(
                    outlet=outlet,
                    item__platform=platform,
                    status_locked=True,
                    is_active_in_outlet=True
                ).select_related('item').order_by('item__item_code')
                lock_display = 'BLS Status Lock'
            
            for idx, io in enumerate(item_outlets, 1):
                item = io.item
                selling_price = io.outlet_selling_price or item.selling_price
                
                rows.append([
                    idx,
                    item.item_code,
                    item.units,
                    item.description[:50] + '...' if len(item.description) > 50 else item.description,
                    float(selling_price) if selling_price else '-',
                    item.pack_description[:30] + '...' if item.pack_description and len(item.pack_description) > 30 else (item.pack_description or '-'),
                    lock_display,
                    f'Branch ({outlet.name})',
                    outlet.name
                ])
        
        return JsonResponse({
            'success': True,
            'headers': headers,
            'rows': rows,
            'total_rows': len(rows),
            'lock_type': lock_type,
            'platform': platform,
            'outlet_name': outlet.name if outlet_id and 'outlet' in locals() else None
        })
        
    except Exception as e:
        logger.error(f"Locked products data API error: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error loading locked products data: {str(e)}'
        })


@login_required
def export_locked_products_api(request):
    """
    Export locked products data to CSV/Excel.
    
    POST params:
    - platform: 'pasons', 'talabat' (required)
    - lock_type: 'cls_price', 'cls_status', 'bls_price', 'bls_status'
    - outlet: outlet_id (optional, for BLS reports)
    - format: 'csv', 'excel'
    """
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from .models import Item, Outlet, ItemOutlet
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required'})
    
    try:
        platform = request.POST.get('platform', '').strip()
        lock_type = request.POST.get('lock_type', '').strip()
        outlet_id = request.POST.get('outlet', '').strip()
        export_format = request.POST.get('format', 'csv').strip()
        
        # Validate platform
        if not platform or platform not in ('pasons', 'talabat'):
            return JsonResponse({'success': False, 'message': 'Valid platform is required'})
        
        # Validate lock type
        if not lock_type or lock_type not in ('cls_price', 'cls_status', 'bls_price', 'bls_status'):
            return JsonResponse({'success': False, 'message': 'Valid lock type is required'})
        
        # Generate filename
        timestamp = timezone.localtime().strftime('%Y-%m-%d-%H%M%S')
        lock_type_name = {
            'cls_price': 'CLS-Price-Locked',
            'cls_status': 'CLS-Status-Locked',
            'bls_price': 'BLS-Price-Locked',
            'bls_status': 'BLS-Status-Locked'
        }[lock_type]
        
        platform_name = platform.title()
        filename = f'{platform_name}-{lock_type_name}-Products-{timestamp}'
        
        if export_format == 'excel':
            # Excel export
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Locked Products'
            
            # Headers
            headers = ['Item Code', 'Units', 'Description', 'Selling Price', 'Pack Description', 'Lock Type', 'Lock Level']
            if lock_type.startswith('bls') and outlet_id:
                headers.append('Outlet')
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            row_num = 2
            if lock_type.startswith('cls'):
                # CLS data
                if lock_type == 'cls_price':
                    items = Item.objects.filter(platform=platform, price_locked=True, is_active=True).order_by('item_code')
                    lock_display = 'CLS Price Lock'
                else:
                    items = Item.objects.filter(platform=platform, status_locked=True, is_active=True).order_by('item_code')
                    lock_display = 'CLS Status Lock'
                
                for item in items:
                    ws.cell(row=row_num, column=1, value=item.item_code)
                    ws.cell(row=row_num, column=2, value=item.units)
                    ws.cell(row=row_num, column=3, value=item.description)
                    ws.cell(row=row_num, column=4, value=float(item.selling_price) if item.selling_price else 0)
                    ws.cell(row=row_num, column=5, value=item.pack_description or '')
                    ws.cell(row=row_num, column=6, value=lock_display)
                    ws.cell(row=row_num, column=7, value='Central (All Outlets)')
                    row_num += 1
            else:
                # BLS data
                if not outlet_id:
                    return JsonResponse({'success': False, 'message': 'Outlet is required for BLS reports'})
                
                outlet = Outlet.objects.get(id=outlet_id, platforms=platform, is_active=True)
                
                if lock_type == 'bls_price':
                    item_outlets = ItemOutlet.objects.filter(
                        outlet=outlet, item__platform=platform, price_locked=True, is_active_in_outlet=True
                    ).select_related('item').order_by('item__item_code')
                    lock_display = 'BLS Price Lock'
                else:
                    item_outlets = ItemOutlet.objects.filter(
                        outlet=outlet, item__platform=platform, status_locked=True, is_active_in_outlet=True
                    ).select_related('item').order_by('item__item_code')
                    lock_display = 'BLS Status Lock'
                
                for io in item_outlets:
                    item = io.item
                    selling_price = io.outlet_selling_price or item.selling_price
                    
                    ws.cell(row=row_num, column=1, value=item.item_code)
                    ws.cell(row=row_num, column=2, value=item.units)
                    ws.cell(row=row_num, column=3, value=item.description)
                    ws.cell(row=row_num, column=4, value=float(selling_price) if selling_price else 0)
                    ws.cell(row=row_num, column=5, value=item.pack_description or '')
                    ws.cell(row=row_num, column=6, value=lock_display)
                    ws.cell(row=row_num, column=7, value=f'Branch ({outlet.name})')
                    ws.cell(row=row_num, column=8, value=outlet.name)
                    row_num += 1
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            wb.save(response)
            return response
        
        else:
            # CSV export
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            
            writer = csv.writer(response)
            
            # Headers
            headers = ['Item Code', 'Units', 'Description', 'Selling Price', 'Pack Description', 'Lock Type', 'Lock Level']
            if lock_type.startswith('bls') and outlet_id:
                headers.append('Outlet')
            writer.writerow(headers)
            
            # Data
            if lock_type.startswith('cls'):
                # CLS data
                if lock_type == 'cls_price':
                    items = Item.objects.filter(platform=platform, price_locked=True, is_active=True).order_by('item_code')
                    lock_display = 'CLS Price Lock'
                else:
                    items = Item.objects.filter(platform=platform, status_locked=True, is_active=True).order_by('item_code')
                    lock_display = 'CLS Status Lock'
                
                for item in items:
                    writer.writerow([
                        item.item_code,
                        item.units,
                        item.description,
                        float(item.selling_price) if item.selling_price else '',
                        item.pack_description or '',
                        lock_display,
                        'Central (All Outlets)'
                    ])
            else:
                # BLS data
                if not outlet_id:
                    return JsonResponse({'success': False, 'message': 'Outlet is required for BLS reports'})
                
                outlet = Outlet.objects.get(id=outlet_id, platforms=platform, is_active=True)
                
                if lock_type == 'bls_price':
                    item_outlets = ItemOutlet.objects.filter(
                        outlet=outlet, item__platform=platform, price_locked=True, is_active_in_outlet=True
                    ).select_related('item').order_by('item__item_code')
                    lock_display = 'BLS Price Lock'
                else:
                    item_outlets = ItemOutlet.objects.filter(
                        outlet=outlet, item__platform=platform, status_locked=True, is_active_in_outlet=True
                    ).select_related('item').order_by('item__item_code')
                    lock_display = 'BLS Status Lock'
                
                for io in item_outlets:
                    item = io.item
                    selling_price = io.outlet_selling_price or item.selling_price
                    
                    writer.writerow([
                        item.item_code,
                        item.units,
                        item.description,
                        float(selling_price) if selling_price else '',
                        item.pack_description or '',
                        lock_display,
                        f'Branch ({outlet.name})',
                        outlet.name
                    ])
            
            return response
        
    except Exception as e:
        logger.error(f"Export locked products error: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Export failed: {str(e)}'})
