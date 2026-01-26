#!/usr/bin/env python3
"""
clean_erp_mixed_columns.py

Take the ERP export (Excel *.xlsx* or CSV) that has the mixed‑up headers
and produce a tidy sheet with the exact order:

item_code | units | promo_price
"""

import argparse
import sys
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# Suppress warnings to keep console clean
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# -------------------------------------------------
# Helper: column letter → 1‑based index (for autofit)
# -------------------------------------------------
def column_index(letter: str) -> int:
    """Convert Excel column letter (A=1, B=2, …) to a numeric index."""
    return sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(letter.upper())))


# -------------------------------------------------
# Helper: auto‑size columns safely
# -------------------------------------------------
def autofit_columns(ws, start_col: str = "A", end_col: str = "E"):
    """Resize each column from start_col to end_col; empty columns get width 8."""
    start_idx = column_index(start_col)
    end_idx   = column_index(end_col)

    for col_idx in range(start_idx, end_idx + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        lengths = [len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None]
        max_len = max(lengths) if lengths else 8
        ws.column_dimensions[col_letter].width = max(max_len + 2, 8)


# -------------------------------------------------
# Core cleaning function
# -------------------------------------------------
def clean_erp_mixed(
    input_path: Path,
    output_path: Path,
    sheet_name: str = None,
) -> None:
    """
    1️⃣ Load the file (Excel or CSV).
    2️⃣ Normalise header names (strip spaces, lower‑case).
    3️⃣ Drop the truly useless columns (SrNo, original Name, Ofr., Price).
    4️⃣ Re‑map the remaining columns to the correct logical names.
    5️⃣ Write a clean Excel file and auto‑fit columns.
    """
    # -------------------------------------------------
    # 1️⃣ Load data (detect file type by suffix)
    # -------------------------------------------------
    if input_path.suffix.lower() in {".csv", ".txt"}:
        df = pd.read_csv(input_path, dtype=str)   # keep everything as string
    else:  # Excel
        raw = pd.read_excel(input_path, sheet_name=sheet_name, engine="openpyxl")
        if isinstance(raw, dict):                 # multiple sheets → use the first
            sheet_used = list(raw.keys())[0]
            df = raw[sheet_used]
        else:
            df = raw
            sheet_used = sheet_name or "default"

    # -------------------------------------------------
    # 2️⃣ Normalise header names (strip spaces, lower‑case)
    # -------------------------------------------------
    original_cols = df.columns.tolist()
    clean_to_orig = {col.strip().lower(): col for col in original_cols}

    # -------------------------------------------------
    # 3️⃣ Drop columns we never need
    # -------------------------------------------------
    unwanted = ["srno", "name", "ofr.", "price"]   # original numeric Name & empty price column
    cols_to_drop = [clean_to_orig[c] for c in unwanted if c in clean_to_orig]
    df.drop(columns=cols_to_drop, inplace=True, errors="ignore")

    # -------------------------------------------------
    # 4️⃣ Build the correct mapping
    # -------------------------------------------------
    # Desired final order (removed Name and Sell.Pr)
    final_order_clean = ["item code", "quantity", "sell.pr"]
    # Map from the *actual* source column to the logical name we need
    mapping = {
        "item code": "item_code",          # keep as‑is
        "quantity":  "units",              # original Quantity → units
        "sell.pr":   "promo_price",        # original Sell.Pr → promo_price
    }

    # Verify that every required logical column exists in the source
    missing = [c for c in final_order_clean if c not in clean_to_orig]
    if missing:
        raise KeyError(
            f"The following expected columns are missing in the source file: {missing}\n"
            f"Available columns (cleaned): {list(clean_to_orig.keys())}"
        )

    # Re‑order and rename
    ordered_original = [clean_to_orig[c] for c in final_order_clean]
    df = df[ordered_original]                     # correct column order
    df.columns = [mapping[c] for c in final_order_clean]  # rename to final titles

    # -------------------------------------------------
    # 5️⃣ Write cleaned workbook
    # -------------------------------------------------
    if output_path.suffix.lower() == ".csv":
        df.to_csv(output_path, index=False)
    else:
        df.to_excel(output_path, index=False, sheet_name="Cleaned", engine="openpyxl")

        # -------------------------------------------------
        # 6️⃣ Auto‑fit columns (Excel only)
        # -------------------------------------------------
        wb = load_workbook(output_path)
        ws = wb["Cleaned"]
        last_col = ws.cell(row=1, column=ws.max_column).column_letter
        autofit_columns(ws, start_col="A", end_col=last_col)
        ws.sheet_view.topLeftCell = "A1"
        wb.save(output_path)


# -------------------------------------------------
# CLI argument handling
# -------------------------------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Re‑order and rename ERP columns that are shifted. "
            "Keeps Item Code, Barcode→Name, Quantity→Unit, Unit→Sell.Pr, "
            "Sell.Pr→Ofr. Price, and drops SrNo and the original numeric Name."
        )
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        default="PromoRaw.xlsx",
        help="Path to the raw file (Excel *.xlsx* or CSV). Default: PromoRaw.xlsx",
    )
    parser.add_argument(
        "output_file",
        nargs="?",
        default=None,
        help="Path for the cleaned file. If omitted, a file named 'cleaned_<input name>.xlsx' is created.",
    )
    parser.add_argument(
        "sheet_name",
        nargs="?",
        default=None,
        help="Sheet name for Excel files (optional). Ignored for CSV.",
    )
    return parser


if __name__ == "__main__":
    parser = build_parser()
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.is_file():
        print(f"❌ Input file not found: {input_path}")
        sys.exit(1)

    # Determine output path
    output_path = (
        Path(args.output_file)
        if args.output_file
        else input_path.parent / f"cleaned_{input_path.stem}.xlsx"
    )

    clean_erp_mixed(input_path, output_path, sheet_name=args.sheet_name)