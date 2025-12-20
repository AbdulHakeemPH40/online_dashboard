"""
Promotion Price Update API Views
Separate file for promotion-related endpoints
"""

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
import json
import logging
import csv
import io

from .promotion_service import PromotionService
from .models import Item, ItemOutlet, Outlet
from .utils import decode_csv_upload

logger = logging.getLogger(__name__)


@login_required
def promotion_update(request):
    """
    Promotion price update page view
    """
    return render(request, 'promotion_update.html', {
        'page_title': 'Promotion Price Update',
        'platforms': ['pasons', 'talabat']
    })


@login_required
def promotion_search_api(request):
    """
    API endpoint to search for item by code and units
    GET /api/promotion/search/?item_code=X&units=Y&platform=Z
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'GET request required'})
    
    item_code = request.GET.get('item_code', '').strip()
    units = request.GET.get('units', '').strip()
    platform = request.GET.get('platform', '').strip()
    
    if not all([item_code, units, platform]):
        return JsonResponse({
            'success': False,
            'message': 'item_code, units, and platform are required'
        })
    
    item_data = PromotionService.search_item(item_code, units, platform)
    
    if not item_data:
        return JsonResponse({
            'success': False,
            'message': f'Item not found: {item_code} ({units}) on {platform}'
        })
    
    return JsonResponse({
        'success': True,
        'item': item_data
    })


@login_required
def promotion_calculate_api(request):
    """
    API endpoint to calculate promotional price preview
    POST /api/promotion/calculate/
    Body: {item_code, units, platform, promo_price, start_date, end_date}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST request required'})
    
    try:
        data = json.loads(request.body)
        
        item_code = data.get('item_code', '').strip()
        units = data.get('units', '').strip()
        platform = data.get('platform', '').strip()
        promo_price = Decimal(str(data.get('promo_price', 0)))
        
        # Get item details
        item_data = PromotionService.search_item(item_code, units, platform)
        
        if not item_data:
            return JsonResponse({
                'success': False,
                'message': 'Item not found'
            })
        
        # Calculate promotional price
        calculation = PromotionService.calculate_promo_price(
            promo_price=promo_price,
            platform=platform,
            item_code=item_code,
            wdf=item_data['wdf'],
            talabat_margin=item_data['talabat_margin'],
            cost=item_data['current_cost'] or Decimal('0'),
            selling_price=item_data['current_selling_price'] or Decimal('0')
        )
        
        return JsonResponse({
            'success': True,
            'calculation': {
                'promo_price': str(calculation['promo_price']),
                'converted_promo': str(calculation['converted_promo']),
                'promo_adjusted': calculation['promo_adjusted'],
                'margin_warning': calculation['margin_warning'],
                'selling_price': str(calculation['selling_price']),
                'selling_adjusted': calculation['selling_adjusted'],
                'selling_warning': calculation['selling_warning'],
                'margin_pct': str(calculation['margin_pct']),
                'difference': str(calculation['difference']),
                'is_wrap': calculation['is_wrap'],
                'wdf': str(calculation['wdf']) if calculation['wdf'] else None
            }
        })
    
    except (ValueError, InvalidOperation) as e:
        return JsonResponse({
            'success': False,
            'message': f'Invalid number format: {str(e)}'
        })
    except Exception as e:
        logger.error(f"Promotion calculate error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error calculating promotion: {str(e)}'
        })


@login_required
def promotion_save_api(request):
    """
    API endpoint to save promotion
    POST /api/promotion/save/
    Body: {item_code, units, platform, promo_price, converted_promo, adjusted_selling, start_date, end_date}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST request required'})
    
    try:
        data = json.loads(request.body)
        
        item_code = data.get('item_code', '').strip()
        units = data.get('units', '').strip()
        platform = data.get('platform', '').strip()
        outlet_id = int(data.get('outlet_id', 0))
        promo_price = Decimal(str(data.get('promo_price', 0)))
        converted_promo = Decimal(str(data.get('converted_promo', 0)))
        adjusted_selling = Decimal(str(data.get('adjusted_selling', 0)))
        start_date_str = data.get('start_date', '')
        end_date_str = data.get('end_date', '')
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Validate dates
        if end_date < start_date:
            return JsonResponse({
                'success': False,
                'message': 'End date must be after start date'
            })
        
        if start_date < date.today():
            return JsonResponse({
                'success': False,
                'message': 'Start date cannot be in the past'
            })
        
        # Save promotion for selected outlet only
        result = PromotionService.save_promotion(
            item_code=item_code,
            units=units,
            platform=platform,
            outlet_id=outlet_id,
            promo_price=promo_price,
            converted_promo=converted_promo,
            adjusted_selling=adjusted_selling,
            start_date=start_date,
            end_date=end_date
        )
        
        return JsonResponse(result)
    
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'message': f'Invalid date or number format: {str(e)}'
        })
    except Exception as e:
        logger.error(f"Promotion save error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error saving promotion: {str(e)}'
        })


@login_required
def promotion_active_api(request):
    """
    API endpoint to get list of active promotions with pagination
    GET /api/promotion/active/?platform=X&outlet=Y&page=1&page_size=20
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'GET request required'})
    
    platform = request.GET.get('platform')
    outlet_id = request.GET.get('outlet')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    # Convert outlet_id to int if provided
    outlet_id = int(outlet_id) if outlet_id else None
    
    result = PromotionService.get_active_promotions(
        platform=platform,
        outlet_id=outlet_id,
        page=page,
        page_size=page_size
    )
    
    return JsonResponse({
        'success': True,
        **result
    })


@login_required
def promotion_bulk_cancel_api(request):
    """
    API endpoint to cancel multiple promotions at once
    POST /api/promotion/bulk-cancel/
    Body: {ids: [1, 2, 3]}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST request required'})
    
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        if not ids:
            return JsonResponse({'success': False, 'message': 'No promotion IDs provided'})
        
        result = PromotionService.bulk_cancel_promotions(ids)
        return JsonResponse(result)
    
    except Exception as e:
        logger.error(f"Bulk cancel error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def promotion_cancel_all_api(request):
    """
    API endpoint to cancel ALL promotions for an outlet
    POST /api/promotion/cancel-all/
    Body: {platform: 'pasons', outlet_id: 1}
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST request required'})
    
    try:
        data = json.loads(request.body)
        platform = data.get('platform')
        outlet_id = data.get('outlet_id')
        
        if not platform or not outlet_id:
            return JsonResponse({'success': False, 'message': 'Platform and outlet_id required'})
        
        result = PromotionService.cancel_all_promotions_for_outlet(platform, int(outlet_id))
        return JsonResponse(result)
    
    except Exception as e:
        logger.error(f"Cancel all error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def promotion_cancel_api(request, promo_id):
    """
    API endpoint to cancel a promotion
    DELETE /api/promotion/{id}/cancel/
    """
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'message': 'DELETE request required'})
    
    result = PromotionService.cancel_promotion(promo_id)
    
    return JsonResponse(result)


@login_required
def bulk_promotion_update(request):
    """
    Bulk promotion update page view
    """
    return render(request, 'bulk_promotion_update.html', {
        'page_title': 'Bulk Promotion Update'
    })


@login_required
def bulk_promotion_preview_api(request):
    """
    Preview bulk promotion CSV upload
    POST /api/bulk-promotion/preview/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST request required'})
    
    try:
        platform = request.POST.get('platform')
        outlet_id = request.POST.get('outlet')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        csv_file = request.FILES.get('csv_file')
        
        if not all([platform, outlet_id, start_date_str, end_date_str, csv_file]):
            return JsonResponse({
                'success': False,
                'message': 'All fields are required'
            })
        
        # Parse datetime (format: yyyy-mm-dd hh:mm:ss) and make timezone-aware
        start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d %H:%M:%S')
        end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S')
        # Make timezone-aware using local timezone (Asia/Dubai from settings)
        start_date = timezone.make_aware(start_date_naive)
        end_date = timezone.make_aware(end_date_naive)
        
        # Decode CSV
        csv_text, encoding = decode_csv_upload(csv_file)
        csv_reader = csv.DictReader(io.StringIO(csv_text))
        csv_data = list(csv_reader)
        
        # Import normalize_csv_header for BOM handling
        from .utils import normalize_csv_header
        
        preview_items = []
        errors = []
        warnings = []
        valid_count = 0
        
        for idx, row in enumerate(csv_data, start=2):
            # Normalize keys to handle BOM/invisible chars
            row = {normalize_csv_header(k): v.strip() if v else '' for k, v in row.items()}
            item_code = row.get('item_code', '')
            units = row.get('units', '')
            promo_price_str = row.get('promo_price', '')
            
            if not all([item_code, units, promo_price_str]):
                errors.append(f"Row {idx}: Missing required fields")
                continue
            
            try:
                promo_price = Decimal(promo_price_str)
            except (ValueError, InvalidOperation):
                errors.append(f"Row {idx}: Invalid promo_price '{promo_price_str}'")
                continue
            
            # Search for item
            item_data = PromotionService.search_item(item_code, units, platform)
            
            if not item_data:
                errors.append(f"Row {idx}: Item not found - {item_code} ({units})")
                continue
            
            # Calculate promotion
            calculation = PromotionService.calculate_promo_price(
                promo_price=promo_price,
                platform=platform,
                item_code=item_code,
                wdf=item_data['wdf'],
                talabat_margin=item_data['talabat_margin'],
                cost=item_data['current_cost'] or Decimal('0'),
                selling_price=item_data['current_selling_price'] or Decimal('0')
            )
            
            status = 'valid'
            if calculation['promo_adjusted'] or calculation['selling_adjusted']:
                status = 'adjusted'
                if calculation['margin_warning']:
                    warnings.append(f"Row {idx}: {calculation['margin_warning']}")
                if calculation['selling_warning']:
                    warnings.append(f"Row {idx}: {calculation['selling_warning']}")
            
            preview_items.append({
                'item_code': item_code,
                'units': units,
                'promo_price': str(promo_price),
                'converted_promo': str(calculation['converted_promo']),
                'selling_price': str(calculation['selling_price']),
                'margin_pct': str(calculation['margin_pct']),
                'status': status
            })
            valid_count += 1
        
        return JsonResponse({
            'success': True,
            'total_items': len(csv_data),
            'valid_items': valid_count,
            'error_count': len(errors),
            'warning_count': len(warnings),
            'errors': errors,
            'warnings': warnings,
            'preview': preview_items[:50]  # Limit preview to 50 items
        })
    
    except Exception as e:
        logger.error(f"Bulk promotion preview error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error previewing CSV: {str(e)}'
        })


@login_required
def bulk_promotion_upload_api(request):
    """
    Apply bulk promotion CSV upload
    POST /api/bulk-promotion/upload/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST request required'})
    
    try:
        platform = request.POST.get('platform')
        outlet_id = request.POST.get('outlet')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        csv_file = request.FILES.get('csv_file')
        
        if not all([platform, outlet_id, start_date_str, end_date_str, csv_file]):
            return JsonResponse({
                'success': False,
                'message': 'All fields are required'
            })
        
        # Parse datetime (format: yyyy-mm-dd hh:mm:ss) and make timezone-aware
        start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d %H:%M:%S')
        end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S')
        # Make timezone-aware using local timezone (Asia/Dubai from settings)
        start_date = timezone.make_aware(start_date_naive)
        end_date = timezone.make_aware(end_date_naive)
        
        # Decode CSV
        csv_text, encoding = decode_csv_upload(csv_file)
        csv_reader = csv.DictReader(io.StringIO(csv_text))
        csv_data = list(csv_reader)
        
        # Import normalize_csv_header for BOM handling
        from .utils import normalize_csv_header
        
        success_count = 0
        error_count = 0
        
        duplicate_count = 0
        updated_count = 0
        
        for row in csv_data:
            # Normalize keys to handle BOM/invisible chars
            row = {normalize_csv_header(k): v.strip() if v else '' for k, v in row.items()}
            item_code = row.get('item_code', '')
            units = row.get('units', '')
            promo_price_str = row.get('promo_price', '')
            
            if not all([item_code, units, promo_price_str]):
                error_count += 1
                continue
            
            try:
                promo_price = Decimal(promo_price_str)
            except (ValueError, InvalidOperation):
                error_count += 1
                continue
            
            # Search for item
            item_data = PromotionService.search_item(item_code, units, platform)
            
            if not item_data:
                error_count += 1
                continue
            
            # Check for existing active promotion
            existing = PromotionService.check_existing_promotion(
                item_code=item_code,
                units=units,
                platform=platform,
                outlet_id=int(outlet_id)
            )
            
            is_update = existing.get('exists', False)
            
            # Calculate promotion
            calculation = PromotionService.calculate_promo_price(
                promo_price=promo_price,
                platform=platform,
                item_code=item_code,
                wdf=item_data['wdf'],
                talabat_margin=item_data['talabat_margin'],
                cost=item_data['current_cost'] or Decimal('0'),
                selling_price=item_data['current_selling_price'] or Decimal('0')
            )
            
            # Save promotion for selected outlet only (will update if exists)
            result = PromotionService.save_promotion(
                item_code=item_code,
                units=units,
                platform=platform,
                outlet_id=int(outlet_id),
                promo_price=promo_price,
                converted_promo=calculation['converted_promo'],
                adjusted_selling=calculation['selling_price'],
                start_date=start_date,
                end_date=end_date
            )
            
            if result['success']:
                if is_update:
                    updated_count += 1
                else:
                    success_count += 1
            else:
                error_count += 1
        
        # Build response message
        msg_parts = []
        if success_count > 0:
            msg_parts.append(f'{success_count} new promotion(s)')
        if updated_count > 0:
            msg_parts.append(f'{updated_count} updated')
        if error_count > 0:
            msg_parts.append(f'{error_count} error(s)')
        
        # Save upload history for tracking
        from .models import UploadHistory, Outlet
        total_records = success_count + updated_count + error_count
        upload_status = 'success' if error_count == 0 else ('partial' if (success_count + updated_count) > 0 else 'failed')
        
        outlet_obj = Outlet.objects.filter(id=outlet_id).first()
        
        UploadHistory.objects.create(
            file_name=csv_file.name,
            platform=platform,
            outlet=outlet_obj,
            update_type='promotion_update',
            records_total=total_records,
            records_success=success_count + updated_count,
            records_failed=error_count,
            records_skipped=0,
            status=upload_status,
            uploaded_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Applied: ' + ', '.join(msg_parts) if msg_parts else 'No changes made',
            'success_count': success_count,
            'updated_count': updated_count,
            'error_count': error_count
        })
    
    except Exception as e:
        logger.error(f"Bulk promotion upload error: {e}", exc_info=True)
        
        # Save failed upload history
        from .models import UploadHistory
        try:
            file_name = csv_file.name if 'csv_file' in locals() and csv_file else 'unknown.csv'
            UploadHistory.objects.create(
                file_name=file_name,
                platform=platform if 'platform' in locals() else 'unknown',
                outlet=None,
                update_type='promotion_update',
                records_total=0,
                records_success=0,
                records_failed=0,
                records_skipped=0,
                status='failed',
                uploaded_by=request.user if request and hasattr(request, 'user') else None,
                error_message=str(e)
            )
        except:
            pass
        
        return JsonResponse({
            'success': False,
            'message': f'Error applying promotions: {str(e)}'
        })


@login_required
def promotion_export_api(request):
    """
    Export promotion items feed (Pasons/Talabat format)
    GET /api/promotion/export/?platform=X&outlet=Y
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'GET request required'})
    
    try:
        platform = request.GET.get('platform')
        outlet_id = request.GET.get('outlet')
        
        if not platform or not outlet_id:
            return JsonResponse({
                'success': False,
                'message': 'Platform and outlet are required'
            })
        
        # Get promotion items
        promo_items = ItemOutlet.objects.filter(
            item__platform=platform,
            outlet_id=outlet_id,
            is_on_promotion=True
        ).select_related('item', 'outlet')
        
        # Get outlet name for filename
        outlet = Outlet.objects.filter(id=outlet_id).first()
        outlet_name = outlet.name.replace(' ', '-').replace('_', '-') if outlet else 'outlet'
        
        # Create filename with outlet name, date and time (using hyphens only)
        from datetime import datetime as dt
        timestamp = dt.now().strftime('%Y-%m-%d-%H-%M-%S')
        filename = f"promo-export-{platform}-{outlet_name}-{timestamp}.csv"
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        if platform == 'pasons':
            # Pasons format: SKU, Promo_price only
            # wrap=9900: use converted_promo (promo_price / wdf)
            # wrap=10000: use promo_price as is
            # Handle multiple SKUs (comma-separated) - export each SKU
            writer.writerow(['SKU', 'Promo_price'])
            
            for io in promo_items:
                # Use converted_promo for wrap=9900, promo_price for wrap=10000
                if io.item.wrap == '9900' and io.converted_promo:
                    promo_price = io.converted_promo
                else:
                    promo_price = io.promo_price or 0
                
                # Handle multiple SKUs (comma-separated)
                skus = io.item.sku.split(',') if io.item.sku else [f"{io.item.item_code}-{io.item.units}"]
                
                for sku in skus:
                    sku = sku.strip()
                    if sku:
                        writer.writerow([sku, promo_price])
        
        else:  # talabat
            # Talabat format: barcode, sku, reason, start_date, end_date, campaign_status, discounted_price, max_no_of_orders, price, active
            writer.writerow(['barcode', 'sku', 'reason', 'start_date', 'end_date', 'campaign_status', 'discounted_price', 'max_no_of_orders', 'price', 'active'])
            
            for io in promo_items:
                # Handle multiple SKUs
                skus = io.item.sku.split(',') if io.item.sku else [f"{io.item.item_code}-{io.item.units}"]
                
                # Get barcode (can be comma-separated like SKU)
                barcodes = io.item.barcode.split(',') if io.item.barcode else ['']
                
                # Ensure barcodes and skus have same length
                if len(barcodes) < len(skus):
                    barcodes.extend([''] * (len(skus) - len(barcodes)))
                
                for idx, sku in enumerate(skus):
                    sku = sku.strip()
                    barcode = barcodes[idx].strip() if idx < len(barcodes) else ''
                    
                    # Format dates as yyyy-mm-dd hh:mm:ss (convert to local timezone)
                    start_date = timezone.localtime(io.promo_start_date).strftime('%Y-%m-%d %H:%M:%S') if io.promo_start_date else ''
                    end_date = timezone.localtime(io.promo_end_date).strftime('%Y-%m-%d %H:%M:%S') if io.promo_end_date else ''
                    
                    # discounted_price = converted_promo (C.Promo)
                    discounted_price = io.converted_promo or 0
                    
                    # price = outlet_selling_price (already the converted selling price for both wrap types)
                    price = io.outlet_selling_price or 0
                    
                    # active = stock status (0 or 1, not text)
                    active = 1 if io.outlet_stock and io.outlet_stock > 0 else 0
                    
                    writer.writerow([
                        barcode,
                        sku,
                        'competitiveness',  # reason
                        start_date,
                        end_date,
                        1,  # campaign_status always 1
                        discounted_price,
                        '',  # max_no_of_orders (empty/null)
                        price,
                        active
                    ])
        
        return response
    
    except Exception as e:
        logger.error(f"Promotion export error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error exporting promotions: {str(e)}'
        })


@login_required
def promotion_erp_export_api(request):
    """
    Export promotion items in ERP format (Talabat only)
    GET /api/promotion/erp-export/?outlet=X
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'GET request required'})
    
    try:
        outlet_id = request.GET.get('outlet')
        
        if not outlet_id:
            return JsonResponse({
                'success': False,
                'message': 'Outlet is required'
            })
        
        # Get Talabat promotion items
        promo_items = ItemOutlet.objects.filter(
            item__platform='talabat',
            outlet_id=outlet_id,
            is_on_promotion=True
        ).select_related('item', 'outlet')
        
        # Get outlet name for filename
        outlet = Outlet.objects.filter(id=outlet_id).first()
        outlet_name = outlet.name.replace(' ', '-').replace('_', '-') if outlet else 'outlet'
        
        # Create filename with outlet name, date and time (using hyphens only)
        from datetime import datetime as dt
        timestamp = dt.now().strftime('%Y-%m-%d-%H-%M-%S')
        filename = f"promo-erp-export-{outlet_name}-{timestamp}.csv"
        
        # Create CSV response with proper csv.writer
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # ERP format: Party, Item Code, Location, Unit, Price
        writer.writerow(['Party', 'Item Code', 'Location', 'Unit', 'Price'])
        
        # Deduplicate: Keep only one entry per (item_code, unit) combination
        # If duplicates exist, keep the one with the LOWEST price (remove largest)
        items_dict = {}  # {(item_code, unit): (party, item_code, location, unit, price)}
        
        for io in promo_items:
            party = 'DT0072'
            item_code = io.item.item_code
            location = ''  # Blank/null
            unit = io.item.units
            
            # Price calculation based on wrap type
            # wrap=9900: C.Promo * WDF (convert back to base price)
            # wrap=10000: C.Promo (use as is)
            if io.item.wrap == '9900' and io.converted_promo and io.item.weight_division_factor and io.item.weight_division_factor > 0:
                price = io.converted_promo * io.item.weight_division_factor
            else:
                price = io.converted_promo or 0
            
            key = (item_code, unit)
            
            # If this key doesn't exist, or if this price is lower than existing, keep it
            if key not in items_dict or price < items_dict[key][4]:
                items_dict[key] = (party, item_code, location, unit, price)
        
        # Write deduplicated rows
        for row in items_dict.values():
            writer.writerow(row)
        
        return response
    
    except Exception as e:
        logger.error(f"Promotion ERP export error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error exporting ERP: {str(e)}'
        })


@login_required
def talabat_promotions_xlsx_export(request):
    """
    Export Talabat active promotions to XLSX format
    GET /api/promotion/talabat-xlsx-export/?outlet=Y
    Exports ALL promotions (no pagination limit)
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'GET request required'})
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        outlet_id = request.GET.get('outlet')
        
        if not outlet_id:
            return JsonResponse({
                'success': False,
                'message': 'Outlet is required'
            })
        
        # Get ALL Talabat promotions for this outlet (no pagination)
        promotions = ItemOutlet.objects.filter(
            is_on_promotion=True,
            item__platform='talabat',
            outlet_id=outlet_id
        ).select_related('item', 'outlet').order_by('-promo_start_date')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Talabat Promotions"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Item Code', 'Description', 'Pack Description', 'Units', 'SKU', 'WDF', 'OCQ', 'MRP', 'Selling', 'Cost', 'C.Cost', 'Promo', 'C.Promo', 'GP %', 'Var', 'Stock', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Define fill colors for conditional formatting
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
        
        # Data rows
        for row_num, io in enumerate(promotions, 2):
            item = io.item
            
            # Calculate converted cost
            is_wrap_9900 = item.wrap == '9900'
            outlet_cost = io.outlet_cost
            
            if is_wrap_9900 and outlet_cost and item.weight_division_factor and item.weight_division_factor > 0:
                converted_cost = float((outlet_cost / item.weight_division_factor).quantize(Decimal('0.01')))
            else:
                converted_cost = float(outlet_cost) if outlet_cost else 0
            
            # Calculate GP% and Variance
            c_promo = float(io.converted_promo) if io.converted_promo else 0
            selling = float(io.outlet_selling_price) if io.outlet_selling_price else 0
            
            # GP% = ((C.Promo - C.Cost) / C.Promo) * 100
            if c_promo > 0:
                gp_percent = ((c_promo - converted_cost) / c_promo) * 100
            else:
                gp_percent = 0
            
            # Variance = Selling - C.Promo
            variance = selling - c_promo
            
            row_data = [
                item.item_code,
                item.description or '',
                item.pack_description or '',
                item.units,
                item.sku,
                float(item.weight_division_factor) if item.weight_division_factor else 0,
                item.outer_case_quantity or 0,
                float(io.outlet_mrp) if io.outlet_mrp else 0,
                selling,
                float(io.outlet_cost) if io.outlet_cost else 0,
                converted_cost,
                float(io.promo_price) if io.promo_price else 0,
                c_promo,
                round(gp_percent, 2),
                round(variance, 2),
                io.outlet_stock or 0,
                'Active' if io.is_on_promotion else 'Inactive'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                
                # Apply color coding
                # GP% column (14th column)
                if col == 14:
                    if gp_percent >= 20:
                        cell.fill = green_fill
                    else:
                        cell.fill = yellow_fill
                
                # Variance column (15th column)
                if col == 15:
                    if variance < 2:
                        cell.fill = yellow_fill
        
        # Adjust column widths
        column_widths = [12, 30, 20, 8, 14, 8, 8, 10, 10, 10, 10, 10, 10, 10, 10, 8, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Get outlet name for filename
        outlet = Outlet.objects.filter(id=outlet_id).first()
        outlet_name = outlet.name.replace(' ', '_') if outlet else 'outlet'
        filename = f"Talabat_Promotions_{outlet_name}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    except ImportError:
        return JsonResponse({
            'success': False,
            'message': 'openpyxl library not installed. Run: pip install openpyxl'
        })
    except Exception as e:
        logger.error(f"Talabat XLSX export error: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error exporting XLSX: {str(e)}'
        })
